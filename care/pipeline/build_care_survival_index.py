"""UK Care Business Survival Index -- data pipeline.

Parses ONS Business Demography 2024 reference tables (Table 5.2a-5.2e: survival of
newly born enterprises by SIC2007 group, one sheet per birth-year cohort 2019-2023)
to build a 1/2/3/4/5-year survival curve for the care sector (SIC 87 residential
care and SIC 88 social work without accommodation), plus a recent-cohort 1-year
survival trend.

Input (already downloaded in this run, reused as-is):
  care/pipeline/raw/ons_business_demography_2024.xlsx

Source: https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable
Licence: Open Government Licence v3.0.

Usage (from repo root):
    python care/pipeline/build_care_survival_index.py
"""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
SRC = RAW / "ons_business_demography_2024.xlsx"
OUT = HERE.parent / "web" / "src" / "data" / "uk-care-business-survival-index.json"
OUT.parent.mkdir(parents=True, exist_ok=True)

# (sheet name, birth year, max survival horizon available in this release)
COHORT_SHEETS = [
    ("Table 5.2a", 2019),
    ("Table 5.2b", 2020),
    ("Table 5.2c", 2021),
    ("Table 5.2d", 2022),
    ("Table 5.2e", 2023),
]

# (SIC2007 group code, label). Matches the CH care SIC set used elsewhere on this
# site (87100/87200/87300/87900/88100/88990) at the 3-digit ONS group level.
SEGMENTS = [
    ("87", "Residential care activities"),
    ("871", "Residential nursing care activities"),
    ("872", "Residential care for learning disability, mental health & substance misuse"),
    ("873", "Residential care activities for the elderly and disabled"),
    ("879", "Other residential care activities"),
    ("88", "Social work activities without accommodation"),
    ("881", "Social work without accommodation for the elderly and disabled"),
    ("889", "Other social work activities without accommodation (incl. supported living)"),
]

HORIZONS = [1, 2, 3, 4, 5]


def _num(v) -> float | None:
    if v in (None, ":", ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_cohort_sheet(ws) -> dict[str, dict]:
    """Return {sic_code: {births, survivors: {h: n|None}, pct: {h: pct|None}}}."""
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        label = row[0]
        if not label:
            continue
        code = str(label).strip().split(" ")[0].rstrip(":")
        if code not in {s[0] for s in SEGMENTS}:
            continue
        births = _num(row[1])
        survivors = {}
        pct = {}
        # columns: 0=label,1=births, then pairs (survivors,pct) for horizons 1..5
        for h in HORIZONS:
            col = 2 + (h - 1) * 2
            survivors[h] = _num(row[col]) if col < len(row) else None
            pct[h] = _num(row[col + 1]) if col + 1 < len(row) else None
        out[code] = {"births": births, "survivors": survivors, "pct": pct}
    return out


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    cohorts_by_sic: dict[str, list[dict]] = {code: [] for code, _ in SEGMENTS}
    for sheet_name, birth_year in COHORT_SHEETS:
        ws = wb[sheet_name]
        parsed = parse_cohort_sheet(ws)
        for code, _label in SEGMENTS:
            row = parsed.get(code)
            if not row:
                continue
            cohorts_by_sic[code].append({
                "birth_year": birth_year,
                "births": row["births"],
                "survival_pct": {str(h): row["pct"][h] for h in HORIZONS},
            })

    segments_out = []
    for code, label in SEGMENTS:
        segments_out.append({
            "sic_group": code,
            "label": label,
            "cohorts": cohorts_by_sic[code],
        })

    # Headline 5-year survival curve: 2019 birth cohort (the only one with a full
    # 5-year horizon in this release), for SIC 87 and SIC 88 top-level groups, plus
    # a combined "all care" curve built from raw births/survivors (not an average
    # of percentages) so the combined ratio is exact.
    def _cohort_2019(code: str) -> dict | None:
        for c in cohorts_by_sic.get(code, []):
            if c["birth_year"] == 2019:
                return c
        return None

    c87 = _cohort_2019("87")
    c88 = _cohort_2019("88")

    def _curve_from_pct(cohort: dict | None) -> list[dict]:
        if not cohort:
            return []
        curve = [{"year": 0, "survival_pct": 100.0}]
        for h in HORIZONS:
            pct = cohort["survival_pct"].get(str(h))
            if pct is not None:
                curve.append({"year": h, "survival_pct": round(pct, 1)})
        return curve

    combined_curve: list[dict] = [{"year": 0, "survival_pct": 100.0}]
    combined_births = None
    if c87 and c88 and c87["births"] is not None and c88["births"] is not None:
        combined_births = c87["births"] + c88["births"]
        for h in HORIZONS:
            p87 = c87["survival_pct"].get(str(h))
            p88 = c88["survival_pct"].get(str(h))
            if p87 is None or p88 is None:
                continue
            survivors_87 = c87["births"] * p87 / 100
            survivors_88 = c88["births"] * p88 / 100
            combined_pct = round((survivors_87 + survivors_88) / combined_births * 100, 1)
            combined_curve.append({"year": h, "survival_pct": combined_pct})

    # 1-year survival trend across the 5 available birth cohorts (2019-2023), SIC 87 and 88
    one_year_trend = {
        "87": [
            {"birth_year": c["birth_year"], "survival_pct": c["survival_pct"].get("1")}
            for c in cohorts_by_sic["87"]
        ],
        "88": [
            {"birth_year": c["birth_year"], "survival_pct": c["survival_pct"].get("1")}
            for c in cohorts_by_sic["88"]
        ],
    }

    snapshot = {
        "meta": {
            "title": "UK Care Business Survival Index",
            "description": (
                "How long newly incorporated UK care businesses survive, by 1, 2, 3, 4 and 5 "
                "years after formation, built from ONS Business Demography birth-cohort data for "
                "SIC 87 (residential care) and SIC 88 (social work without accommodation)."
            ),
            "generated_at": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "pull_date": dt.date.today().isoformat(),
            "licence": "Open Government Licence v3.0",
            "source": {
                "name": "Business Demography 2024 reference tables",
                "publisher": "Office for National Statistics",
                "url": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable",
                "release_date": "2025-11-20",
                "licence": "Open Government Licence v3.0",
            },
            "attribution": (
                "UK Care Business Survival Index data compiled from ONS Business Demography "
                "reference tables (Open Government Licence v3.0). Free to cite with attribution "
                "to Care Home Tax."
            ),
            "methodology": (
                "Survival rates are ONS's own published enterprise-level birth-cohort survival "
                "figures (Tables 5.2a-5.2e), read directly from the Standard Industrial "
                "Classification 2007 group breakdown: not re-derived or estimated. An 'enterprise' "
                "birth is a new business appearing on the Inter-Departmental Business Register in "
                "a given year; survival is measured against VAT/PAYE registration remaining active "
                "at each anniversary. Each birth-year cohort (2019 to 2023) can only show survival "
                "up to however many years have elapsed since that release: the 2019 cohort has the "
                "full 5-year curve, the 2023 cohort only 1-year. The combined 'all care' curve "
                "sums raw births and survivor counts for SIC 87 and SIC 88 before dividing, so it "
                "is an exact weighted figure, not an average of the two percentages."
            ),
            "caveats": [
                "SIC group counts are enterprise-level, not location-level: a single care company "
                "with multiple registered locations counts once here, unlike the CQC-based UK Care "
                "Home Density & Quality Index, which counts regulated locations.",
                "SIC is self-reported at incorporation; domiciliary care agencies sometimes "
                "register under residential SIC codes (87xxx) rather than 88100, so the SIC 87 / "
                "SIC 88 split is indicative, not exact.",
                "Counts are control-rounded to base 5 by ONS disclosure control, so small "
                "percentage movements between adjacent cohorts can reflect rounding as much as "
                "real change.",
                "'Survival' means the enterprise remained VAT/PAYE-registered at that anniversary; "
                "it does not mean the same care service, ownership, or CQC registration continued "
                "unchanged (a home can survive as a business while changing registered manager or "
                "provider on the CQC register).",
                "Only the 2019 birth cohort has a complete 5-year curve in this release; later "
                "cohorts (2020 to 2023) are shown only as far as ONS has published, so the "
                "1-year-survival trend across cohorts is the fairest recent-years comparison.",
            ],
        },
        "segments": segments_out,
        "headline": {
            "sic_87_5yr_curve": _curve_from_pct(c87),
            "sic_88_5yr_curve": _curve_from_pct(c88),
            "combined_5yr_curve": combined_curve,
            "combined_2019_births": combined_births,
            "one_year_survival_trend": one_year_trend,
        },
    }

    OUT.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")
    print(f"[done] wrote {OUT} ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  SIC 87 5yr curve: {snapshot['headline']['sic_87_5yr_curve']}")
    print(f"  SIC 88 5yr curve: {snapshot['headline']['sic_88_5yr_curve']}")
    print(f"  Combined 5yr curve: {snapshot['headline']['combined_5yr_curve']}")


if __name__ == "__main__":
    main()
