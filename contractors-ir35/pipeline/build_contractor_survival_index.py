"""UK Contractor Survival Index -- data pipeline.

Parses ONS Business Demography 2024 reference tables (Table 5.2a-5.2e: survival of
newly born enterprises by SIC2007 group, one sheet per birth-year cohort 2019-2023)
to build a 1/2/3/4/5-year survival curve for the contractor SIC groups: 62 (computer
programming, consultancy and related activities), 70 (activities of head offices;
management consultancy activities), 71 (architectural and engineering activities;
technical testing and analysis), plus a combined "all contractor SIC groups" curve
and the all-industries baseline ("Total" row in the same tables).

Adapted from care/pipeline/build_care_survival_index.py (same source, same table
layout, different SIC groups).

Input (already downloaded for the care site, reused as-is, not re-downloaded):
  care/pipeline/raw/ons_business_demography_2024.xlsx

Source: https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable
Licence: Open Government Licence v3.0.

Usage (from repo root):
    python contractors-ir35/pipeline/build_contractor_survival_index.py
"""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
RAW = HERE.parent.parent / "care" / "pipeline" / "raw"
SRC = RAW / "ons_business_demography_2024.xlsx"
OUT = HERE.parent / "web" / "src" / "data" / "uk-contractor-survival-index.json"
OUT.parent.mkdir(parents=True, exist_ok=True)

# (sheet name, birth year)
COHORT_SHEETS = [
    ("Table 5.2a", 2019),
    ("Table 5.2b", 2020),
    ("Table 5.2c", 2021),
    ("Table 5.2d", 2022),
    ("Table 5.2e", 2023),
]

# Contractor SIC2007 2-digit groups. Coarser than the exact contractor SIC codes
# used in the UK Contractor Index (62020, 70229, 71121/71122 etc): this table only
# breaks survival down to 2/3-digit group level, so the survival figures cover the
# whole division, including non-contractor firms within it (e.g. games studios
# under 62, PR firms under 70210, portrait photography adjacent activity).
GROUPS = [
    ("62", "Computer programming, consultancy and related activities"),
    ("70", "Activities of head offices; management consultancy activities"),
    ("71", "Architectural and engineering activities; technical testing and analysis"),
]
ALL_INDUSTRIES_LABEL = "Total"
HORIZONS = [1, 2, 3, 4, 5]


def _num(v) -> float | None:
    if v in (None, ":", ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_cohort_sheet(ws) -> dict[str, dict]:
    """Return {code: {births, survivors: {h: n|None}, pct: {h: pct|None}}} for our
    3 contractor groups plus the 'Total' (all-industries) row."""
    wanted = {code for code, _ in GROUPS} | {ALL_INDUSTRIES_LABEL}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        label = row[0]
        if not label:
            continue
        raw = str(label).strip()
        if raw == ALL_INDUSTRIES_LABEL:
            code = ALL_INDUSTRIES_LABEL
        else:
            code = raw.split(" ")[0].rstrip(":")
        if code not in wanted:
            continue
        births = _num(row[1])
        survivors = {}
        pct = {}
        # columns: 0=label, 1=births, then pairs (survivors, pct) for horizons 1..5
        for h in HORIZONS:
            col = 2 + (h - 1) * 2
            survivors[h] = _num(row[col]) if col < len(row) else None
            pct[h] = _num(row[col + 1]) if col + 1 < len(row) else None
        out[code] = {"births": births, "survivors": survivors, "pct": pct}
    return out


def _stats(row: dict | None) -> dict:
    """Shape a parsed row into the SurvivalCohortStats field layout."""
    if row is None:
        return {
            "births": None,
            **{f"y{h}_count": None for h in HORIZONS},
            **{f"y{h}_pct": None for h in HORIZONS},
        }
    out: dict = {"births": row["births"]}
    for h in HORIZONS:
        out[f"y{h}_count"] = row["survivors"].get(h)
        out[f"y{h}_pct"] = round(row["pct"][h], 1) if row["pct"].get(h) is not None else None
    return out


def _combine(rows: list[dict | None]) -> dict | None:
    """Sum raw births/survivors across groups before dividing, so the combined
    percentage is an exact weighted figure, not an average of percentages."""
    present = [r for r in rows if r is not None and r["births"] is not None]
    if not present:
        return None
    births = sum(r["births"] for r in present)
    survivors = {}
    pct = {}
    for h in HORIZONS:
        vals = [r["survivors"].get(h) for r in present]
        if any(v is None for v in vals):
            survivors[h] = None
            pct[h] = None
            continue
        s = sum(vals)
        survivors[h] = s
        pct[h] = round(s / births * 100, 1) if births else None
    return {"births": births, "survivors": survivors, "pct": pct}


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    parsed_by_year: dict[int, dict[str, dict]] = {}
    for sheet_name, birth_year in COHORT_SHEETS:
        parsed_by_year[birth_year] = parse_cohort_sheet(wb[sheet_name])

    # Per-group cohort tables (breakdown, one entry per SIC group)
    groups_out = []
    for code, label in GROUPS:
        cohorts = []
        for _sheet, birth_year in COHORT_SHEETS:
            row = parsed_by_year[birth_year].get(code)
            if not row:
                continue
            cohorts.append({
                "birth_year": birth_year,
                "births": row["births"],
                **{f"y{h}_pct": (round(row["pct"][h], 1) if row["pct"].get(h) is not None else None) for h in HORIZONS},
            })
        groups_out.append({"sic_group": code, "label": label, "cohorts": cohorts})

    # Combined cohorts: "contractor" (62+70+71 summed) vs "all_industries" (Total row)
    cohorts_out = []
    for _sheet, birth_year in COHORT_SHEETS:
        parsed = parsed_by_year[birth_year]
        combined = _combine([parsed.get(code) for code, _ in GROUPS])
        all_ind = parsed.get(ALL_INDUSTRIES_LABEL)
        cohorts_out.append({
            "birth_year": birth_year,
            "contractor": _stats(combined),
            "all_industries": _stats(all_ind),
        })

    # Headline: latest cohort with a full 5-year horizon (2019), and latest cohort
    # with 1-year data (2023, most recent births).
    latest_5yr = next((c for c in cohorts_out if c["contractor"]["y5_pct"] is not None), None)
    latest_1yr = next((c for c in reversed(cohorts_out) if c["contractor"]["y1_pct"] is not None), None)

    snapshot = {
        "meta": {
            "generated_at": dt.date.today().isoformat(),
            "title": "UK Contractor Survival Index",
            "description": (
                "How long UK contractor-sector businesses survive, by 1, 2, 3, 4 and 5 years "
                "after formation, built from ONS Business Demography birth-cohort data for SIC "
                "groups 62 (computer programming and IT consultancy), 70 (head offices and "
                "management consultancy) and 71 (architecture and engineering), against the "
                "all-industries UK average."
            ),
            "source_url": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable",
            "release_page": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/bulletins/businessdemography/2024",
            "release_date": "2025-11-20",
            "publisher": "Office for National Statistics",
            "licence": "Open Government Licence v3.0",
            "sources": [
                {
                    "name": "Business Demography 2024 reference tables (Table 5.2a-5.2e)",
                    "publisher": "Office for National Statistics",
                    "url": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable",
                    "release_page": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/bulletins/businessdemography/2024",
                    "licence": "Open Government Licence v3.0",
                    "retrieved": dt.date.today().isoformat(),
                    "release_date": "2025-11-20",
                    "attribution": (
                        "Data sourced from the Office for National Statistics under the Open "
                        "Government Licence v3.0. Free to cite with attribution to Contractor "
                        "Tax Accountants."
                    ),
                }
            ],
            "attribution": (
                "UK Contractor Survival Index data compiled from ONS Business Demography "
                "reference tables (Open Government Licence v3.0). Free to cite with attribution "
                "to Contractor Tax Accountants."
            ),
            "notes": (
                "Survival rates are ONS's own published enterprise-level birth-cohort survival "
                "figures (Table 5.2a-5.2e), read directly from the Standard Industrial "
                "Classification 2007 group breakdown, not re-derived or estimated. An "
                "'enterprise' birth is a new business appearing on the Inter-Departmental "
                "Business Register in a given year (VAT or PAYE registration); survival is "
                "measured against that registration remaining active at each anniversary. Each "
                "birth-year cohort (2019 to 2023) can only show survival up to however many "
                "years have elapsed since that release: the 2019 cohort has the full 5-year "
                "curve, the 2023 cohort only 1-year. The combined 'contractor' curve sums raw "
                "births and survivor counts for SIC groups 62, 70 and 71 before dividing, so it "
                "is an exact weighted figure, not an average of the three percentages."
            ),
            "caveats": [
                "SIC groups 62, 70 and 71 are broader than the specific SIC codes used in the "
                "UK Contractor Index (62020, 70229, 71121/71122 etc): this ONS table only "
                "breaks survival down to 2-digit (occasionally 3-digit) group level, so figures "
                "here include non-contractor firms within the same group, for example games "
                "studios and software product companies under group 62, PR and communications "
                "firms under group 70, and architecture practices under group 71.",
                "'Enterprise', not 'company': this is a different unit to our UK Contractor "
                "Index and UK Contractor Insolvency Index, which both track Companies House "
                "limited companies specifically. The ONS enterprise measure is broader: it "
                "includes sole traders and partnerships registered for VAT or PAYE as well as "
                "limited companies. The two datasets should not be combined into a single "
                "figure.",
                "Figures are control-rounded to the base 5 by ONS disclosure control, so small "
                "percentage movements between adjacent cohorts can reflect rounding as much as "
                "real change.",
                "'Survival' means the enterprise remained VAT/PAYE-registered at that "
                "anniversary; it does not mean the business is profitable, growing, or trading "
                "under the same name, structure or ownership.",
            ],
        },
        "headline": {
            "latest_5yr_cohort_year": latest_5yr["birth_year"] if latest_5yr else None,
            "latest_5yr_contractor_pct": latest_5yr["contractor"]["y5_pct"] if latest_5yr else None,
            "latest_5yr_all_industries_pct": latest_5yr["all_industries"]["y5_pct"] if latest_5yr else None,
            "latest_1yr_cohort_year": latest_1yr["birth_year"] if latest_1yr else None,
            "latest_1yr_contractor_pct": latest_1yr["contractor"]["y1_pct"] if latest_1yr else None,
            "latest_1yr_all_industries_pct": latest_1yr["all_industries"]["y1_pct"] if latest_1yr else None,
        },
        "cohorts": cohorts_out,
        "groups": groups_out,
    }

    OUT.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")
    print(f"[done] wrote {OUT} ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  Latest 5yr cohort: {snapshot['headline']['latest_5yr_cohort_year']} "
          f"contractor={snapshot['headline']['latest_5yr_contractor_pct']}% "
          f"all_industries={snapshot['headline']['latest_5yr_all_industries_pct']}%")
    print(f"  Latest 1yr cohort: {snapshot['headline']['latest_1yr_cohort_year']} "
          f"contractor={snapshot['headline']['latest_1yr_contractor_pct']}% "
          f"all_industries={snapshot['headline']['latest_1yr_all_industries_pct']}%")

    # Self-check
    assert len(cohorts_out) == 5, f"expected 5 birth-year cohorts, got {len(cohorts_out)}"
    assert len(groups_out) == 3, f"expected 3 SIC groups, got {len(groups_out)}"
    for c in cohorts_out:
        assert c["contractor"]["births"] is not None, f"missing contractor births for {c['birth_year']}"
        assert c["all_industries"]["births"] is not None, f"missing all_industries births for {c['birth_year']}"
    print("Self-check passed.")


if __name__ == "__main__":
    main()
