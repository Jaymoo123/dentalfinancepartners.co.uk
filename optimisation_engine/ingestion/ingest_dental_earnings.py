"""
NHS Dentist Earnings and Expenses Tracker ingester.

Sources (2024/25 edition, published 30 July 2026):
  1. Cross-sectional CSV:
     https://files.digital.nhs.uk/EC/30B458/dentearexp_202425_csv.csv
  2. UK Time Series XLSX:
     https://files.digital.nhs.uk/02/EA6696/Dental%20Earnings%20and%20Expenses%20UK%20Timeseries%20202425.xlsx
  Publication page:
     https://digital.nhs.uk/data-and-information/publications/statistical/dental-earnings-and-expenses-estimates/2024-25
Licence: Open Government Licence v3.0

The ingester:
  - Parses the cross-sectional CSV to extract England/country/region breakdowns
    for self-employed primary-care NHS dentists in the reference year
  - Parses Table 1.1 of the time-series XLSX for the England multi-year trend
  - Writes Dentists/web/src/data/nhs-dental-earnings-index.json

Usage:
    python optimisation_engine/ingestion/ingest_dental_earnings.py

Requires openpyxl. There is deliberately NO hardcoded fallback series: an earlier
version fell back to a hand-keyed table that did not match the published figures
and shipped for months. If the XLSX cannot be parsed the run fails loudly instead.

Self-check assertion at end.
"""

import json
import csv
import io
import sys
import datetime
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

EDITION = "2024/25"
PUBLISHED = "2026-07-30"
CSV_URL = "https://files.digital.nhs.uk/EC/30B458/dentearexp_202425_csv.csv"
XLSX_URL = "https://files.digital.nhs.uk/02/EA6696/Dental%20Earnings%20and%20Expenses%20UK%20Timeseries%20202425.xlsx"
PUBLICATION_PAGE = "https://digital.nhs.uk/data-and-information/publications/statistical/dental-earnings-and-expenses-estimates/2024-25"

OUT_PATH = Path(__file__).parents[2] / "Dentists" / "web" / "src" / "data" / "nhs-dental-earnings-index.json"

# ------------------------------------------------------------------ helpers --

def fetch_url(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "DentalFinancePartners-Research/1.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()


# Every breakdown column in the cross-sectional CSV except Country and Region. A row is
# only a top-level figure when all of these are blank; otherwise it is a slice (by age,
# gender, practice size and so on) that must not be mixed into the headline.
SEGMENTATION_COLUMNS = [
    "Earnings to Expenses Ratio (%)",
    "SIMD",
    "Age_Band",
    "Gender",
    "Activity_Type",
    "Percentage_Gross_Earnings_from_Health_Service_Dentistry",
    "Practice_Size",
    "Weekly_Working_Hours",
    "Percentage_of_Time_Spent_on_NHS_Dentistry",
    "Business_Arrangement",
    "Range of Taxable Income",
]


def fmt_val(v: str) -> float | None:
    """Parse numeric string from CSV; return None for blanks/non-numeric."""
    s = str(v or "").strip().replace(",", "")
    if not s or s in ("-", "...", "na", "n/a", "[c]"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# --------------------------------------------------- cross-sectional parser --

def parse_cross_sectional_csv(content: bytes) -> dict:
    """
    The CSV has columns: Country, Dental_Type, Contract_Type, ..., Measure, VALUE
    We want rows where:
      - Dental_Type in ('All', 'GDS') -- all self-employed primary care
      - Contract_Type == 'All'
      - Business_Arrangement == '' (i.e. top-level, not split by arrangement)
      - Measure in (Average Income Before Tax, Median Income Before Tax,
                    Average Gross Earnings, Average Expenses, ...)

    Returns a dict: { country: { region: { measure: value } } }
    """
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    MEASURES_OF_INTEREST = {
        "Average Gross Earnings",
        "Average Expenses",
        "Average Income Before Tax",
        "Median Income Before Tax",
        "Sample Count",
        "Estimated Population",
    }

    results = defaultdict(lambda: defaultdict(dict))

    for row in rows:
        measure = (row.get("Measure") or "").strip()
        if measure not in MEASURES_OF_INTEREST:
            continue

        dental_type = (row.get("Dental_Type") or "").strip()
        contract_type = (row.get("Contract_Type") or "").strip()

        # Top-level only: All dental types, All contract types, and no sub-segmentation
        # on ANY of the breakdown columns. Region is the one breakdown we keep.
        if dental_type != "All" or contract_type != "All":
            continue
        if any((row.get(col) or "").strip() for col in SEGMENTATION_COLUMNS):
            continue

        country = (row.get("Country") or "").strip() or "England"
        # The 2023/24 edition called these Region_Code and VALUE; 2024/25 renamed them
        # Region and Value. Accept either so an edition bump does not silently empty the
        # regional breakdown.
        region = (row.get("Region") or row.get("Region_Code") or "").strip() or "_national"

        val = fmt_val(row.get("Value") if row.get("Value") is not None else row.get("VALUE", ""))
        if val is not None:
            results[country][region][measure] = val

    return dict(results)


def build_national_snapshot(cross: dict) -> dict:
    """Extract England national row."""
    england = cross.get("England", {})
    national = england.get("_national", {})
    return {
        "country": "England",
        "year": EDITION,
        "avg_gross_earnings": national.get("Average Gross Earnings"),
        "avg_expenses": national.get("Average Expenses"),
        "avg_net_income": national.get("Average Income Before Tax"),
        "median_net_income": national.get("Median Income Before Tax"),
        "sample_count": national.get("Sample Count"),
        "estimated_population": national.get("Estimated Population"),
    }


def build_country_breakdown(cross: dict) -> list[dict]:
    rows = []
    for country, regions in cross.items():
        nat = regions.get("_national", {})
        if not nat:
            continue
        rows.append({
            "country": country,
            "year": EDITION,
            "avg_gross_earnings": nat.get("Average Gross Earnings"),
            "avg_expenses": nat.get("Average Expenses"),
            "avg_net_income": nat.get("Average Income Before Tax"),
            "median_net_income": nat.get("Median Income Before Tax"),
        })
    return sorted(rows, key=lambda r: r["country"])


def build_regional_breakdown(cross: dict) -> list[dict]:
    """England regional breakdown (Region_Code != '_national')."""
    england = cross.get("England", {})
    rows = []
    for region_code, measures in england.items():
        if region_code == "_national":
            continue
        rows.append({
            "region_code": region_code,
            "year": EDITION,
            "avg_gross_earnings": measures.get("Average Gross Earnings"),
            "avg_expenses": measures.get("Average Expenses"),
            "avg_net_income": measures.get("Average Income Before Tax"),
            "median_net_income": measures.get("Median Income Before Tax"),
        })
    return sorted(rows, key=lambda r: r["region_code"])


# --------------------------------------------------- time-series XLSX parser --

# Table 1.1 of the UK time series workbook. Layout: a banner block, then a header row
# whose first cell is "Country", then one row per (Country, Dental Type, Contract Type,
# Category) with one column per financial year. Cells read "z" where the breakdown is
# Not applicable for that country/year (England is "z" for 2008/09 to 2016/17).
TIMESERIES_SHEET = "Table 1.1 Earn - Contract Type"

# Category label in the workbook -> key in our output rows.
TIMESERIES_CATEGORIES = {
    "Gross Earnings": "avg_gross_earnings",
    "Total Expenses": "avg_expenses",
    "Income Before Tax": "avg_net_income",
}


def parse_timeseries_xlsx(content: bytes, country: str = "England") -> list[dict]:
    """
    Parse Table 1.1 for one country, Dental Type = All, Contract Type = All.

    Returns [{year, avg_gross_earnings, avg_expenses, avg_net_income}, ...] for every
    year the workbook actually reports, oldest first. Years marked "z" (Not applicable)
    are omitted rather than guessed at.

    Raises if the sheet, the header or the England/All/All block cannot be found: a
    silently empty series is how the fabricated fallback got shipped in the first place.
    """
    import openpyxl  # hard dependency on purpose, see module docstring
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    if TIMESERIES_SHEET not in wb.sheetnames:
        raise ValueError(f"sheet {TIMESERIES_SHEET!r} not in workbook: {wb.sheetnames}")
    ws = wb[TIMESERIES_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(rows) if r and str(r[0] or "").strip() == "Country"),
        None,
    )
    if header_idx is None:
        raise ValueError("no header row starting 'Country' in Table 1.1")

    header = rows[header_idx]
    # Year columns start after Country / Dental Type / Contract Type / Category and are
    # labelled "2024/25", sometimes with a trailing "[note 26]" on its own line.
    years = {}
    for col, cell in enumerate(header):
        label = str(cell or "").splitlines()[0].strip() if cell else ""
        if col >= 4 and len(label) == 7 and label[4] == "/":
            years[col] = label
    if not years:
        raise ValueError("no year columns found in Table 1.1 header")

    by_year: dict[str, dict] = {label: {"year": label} for label in years.values()}
    seen = 0
    for r in rows[header_idx + 1:]:
        if not r or str(r[0] or "").strip() != country:
            continue
        if str(r[1] or "").strip() != "All" or str(r[2] or "").strip() != "All":
            continue
        key = TIMESERIES_CATEGORIES.get(str(r[3] or "").strip())
        if not key:
            continue
        seen += 1
        for col, label in years.items():
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)):
                by_year[label][key] = round(float(v))

    if seen != len(TIMESERIES_CATEGORIES):
        raise ValueError(
            f"expected {len(TIMESERIES_CATEGORIES)} category rows for {country}/All/All, found {seen}"
        )

    # Keep only years where the workbook reports all three measures.
    series = [
        by_year[label]
        for label in years.values()
        if all(k in by_year[label] for k in TIMESERIES_CATEGORIES.values())
    ]
    series.sort(key=lambda row: row["year"])
    if not series:
        raise ValueError(f"no reported years for {country}/All/All in Table 1.1")
    return series


# ----------------------------------------------------------------------- main --

def main():
    print("Fetching cross-sectional CSV ...", file=sys.stderr)
    csv_content = fetch_url(CSV_URL)
    print(f"  Downloaded {len(csv_content):,} bytes", file=sys.stderr)
    cross_data = parse_cross_sectional_csv(csv_content)

    national = build_national_snapshot(cross_data)
    country_breakdown = build_country_breakdown(cross_data)
    regional_breakdown = build_regional_breakdown(cross_data)

    print("Fetching time-series XLSX ...", file=sys.stderr)
    xlsx_content = fetch_url(XLSX_URL)
    print(f"  Downloaded {len(xlsx_content):,} bytes", file=sys.stderr)
    timeseries = parse_timeseries_xlsx(xlsx_content, country="England")
    print(
        f"  Parsed {len(timeseries)} reported years from Table 1.1 "
        f"({timeseries[0]['year']} to {timeseries[-1]['year']})",
        file=sys.stderr,
    )

    latest_ts = timeseries[-1]
    prev_ts = timeseries[-2] if len(timeseries) >= 2 else None
    net_change_yoy = None
    if prev_ts and prev_ts.get("avg_net_income") and latest_ts.get("avg_net_income"):
        net_change_yoy = round(latest_ts["avg_net_income"] - prev_ts["avg_net_income"])

    snapshot = {
        "meta": {
            "generated_at": datetime.date.today().isoformat(),
            "reference_year": EDITION,
            "edition": f"Dental Earnings and Expenses Estimates, {EDITION}",
            "edition_published": PUBLISHED,
            "timeseries_coverage": f"{timeseries[0]['year']} to {timeseries[-1]['year']}",
            "sources": [
                {
                    "name": f"Dental Earnings and Expenses Estimates, {EDITION} (CSV)",
                    "publisher": "NHS England / NHS Digital",
                    "url": CSV_URL,
                    "publication_page": PUBLICATION_PAGE,
                    "licence": "Open Government Licence v3.0",
                    "retrieved": datetime.date.today().isoformat(),
                    "attribution": "Data sourced from NHS England Digital under the Open Government Licence v3.0. Free to cite with attribution to Dental Finance Partners.",
                },
                {
                    "name": f"Dental Earnings and Expenses UK Timeseries {EDITION} (XLSX)",
                    "publisher": "NHS England / NHS Digital",
                    "url": XLSX_URL,
                    "publication_page": PUBLICATION_PAGE,
                    "licence": "Open Government Licence v3.0",
                    "retrieved": datetime.date.today().isoformat(),
                    "timeseries_source": f"parsed_from_xlsx:{TIMESERIES_SHEET}",
                },
            ],
            "notes": (
                "Figures cover self-employed primary-care NHS dentists. The headline and the "
                "time series are ENGLAND only, taken from Table 1.1 of the UK time series "
                "workbook (Country = England, Dental Type = All, Contract Type = All). The "
                "country table also reports Wales; Scotland and Northern Ireland report on "
                "different contract categories and have no comparable All/All row. "
                "The time series starts at 2017/18 because the workbook marks England "
                "'z Not applicable' for 2008/09 to 2016/17. "
                "Average Income Before Tax = average gross earnings minus average expenses "
                "(before income tax and national insurance). "
                "2020/21 figures are distorted by NHS Covid support payments to dental contractors. "
                "Values are rounded to the nearest hundred pounds as published by NHS England Digital. "
                "Private earnings of dentists who are primarily private are not represented."
            ),
        },
        "headline": {
            "reference_year": EDITION,
            "avg_net_income_england": national.get("avg_net_income"),
            "median_net_income_england": national.get("median_net_income"),
            "avg_gross_earnings_england": national.get("avg_gross_earnings"),
            "avg_expenses_england": national.get("avg_expenses"),
            "estimated_population_england": national.get("estimated_population"),
            "net_income_change_yoy": net_change_yoy,
            "prior_year": prev_ts.get("year") if prev_ts else None,
            "prior_year_avg_net_income": prev_ts.get("avg_net_income") if prev_ts else None,
        },
        "cross_sectional_latest": {
            "national": national,
            "by_country": country_breakdown,
            "by_region_england": regional_breakdown,
        },
        "timeseries_england": timeseries,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Written: {OUT_PATH}", file=sys.stderr)

    # Self-check
    assert national.get("avg_net_income") is not None, "No England national row parsed from CSV"
    assert timeseries, "Empty time series"
    for row in timeseries:
        for key in ("avg_gross_earnings", "avg_expenses", "avg_net_income"):
            assert row.get(key) is not None, f"{row['year']}: missing {key}"
            assert row[key] >= 0, f"{row['year']}: negative {key}"
    # The cross-sectional CSV and the time series are two different files describing the
    # same reference year. If they disagree, one of them was read wrong.
    assert latest_ts["year"] == EDITION, f"time series ends at {latest_ts['year']}, expected {EDITION}"
    for ts_key, cs_key in (
        ("avg_gross_earnings", "avg_gross_earnings"),
        ("avg_expenses", "avg_expenses"),
        ("avg_net_income", "avg_net_income"),
    ):
        assert latest_ts[ts_key] == national[cs_key], (
            f"{EDITION} {ts_key}: XLSX says {latest_ts[ts_key]}, CSV says {national[cs_key]}"
        )
    print("Self-check PASSED", file=sys.stderr)


if __name__ == "__main__":
    main()
