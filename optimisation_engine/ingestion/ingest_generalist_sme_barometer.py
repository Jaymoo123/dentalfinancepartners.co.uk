"""UK Small Business Barometer (generalist flagship) -- data ingestion.

Fuses four national, all-sector series into one barometer for the generalist
(Holloway Davies) research flagship. Unlike the niche sites' single-SIC
indexes, every figure here is ALL-SECTOR / national.

Sources (all Open Government Licence v3.0):
  1. Companies House -- "Incorporated companies in the UK" quarterly official
     statistics CSV (incorporations, dissolutions, register size). UK-wide.
  2. The Insolvency Service -- Company Insolvency Statistics, Industry Tables
     (Table A1a: company insolvencies by 1-digit SIC section, monthly +
     annual). England and Wales only (Scotland/NI published separately).
     The "Total" row is the all-sector national monthly/annual series.
  3. ONS Business Demography -- Table 4.2 (survival of newly-born enterprises
     by broad industry group). The "Total" row (all industries) is the
     founder-survival curve.
  4. ONS/DBT Business Population Estimates -- Table 1 (size breakdown),
     Table 3 (legal status split), Table 25 (UK time series 2010-2025).

Usage:
  python -m optimisation_engine.ingestion.ingest_generalist_sme_barometer

Self-contained: no Supabase, no git commit. Update the four release URLs
below each time a newer quarterly/monthly/annual edition is published.
"""
from __future__ import annotations

import csv
import io
import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import requests

UA = {"User-Agent": "Mozilla/5.0"}
LICENCE = "Open Government Licence v3.0"
ATTRIBUTION = (
    "UK Small Business Barometer compiled from Companies House, The Insolvency "
    "Service and Office for National Statistics public data (Open Government "
    "Licence v3.0). Free to cite with attribution to Holloway Davies."
)

# --- Release pins (update each quarter/month as new editions publish) ------
CH_CSV_URL = (
    "https://assets.publishing.service.gov.uk/media/69f09478b0c3a4023e5d6d56/"
    "Incorporated_companies_in_the_UK_January_to_March_2026.csv"
)
CH_RELEASE_PAGE = "https://www.gov.uk/government/statistics/incorporated-companies-in-the-uk-january-to-march-2026"

INSOLVENCY_INDUSTRY_XLSX_URL = (
    "https://assets.publishing.service.gov.uk/media/6a58e5a15ca06bf11ccb4335/"
    "Industry_Tables_in_Excel__xlsx__Format_-_Company_Insolvency_Statistics_June_2026.xlsx"
)
INSOLVENCY_DATA_XLSX_URL = (
    "https://assets.publishing.service.gov.uk/media/6a57a5272cbe5d1c179a65bb/"
    "Data_Tables_in_Excel__xlsx__Format_-_Company_Insolvency_Statistics_June_2026.xlsx"
)
INSOLVENCY_RELEASE_PAGE = "https://www.gov.uk/government/statistics/company-insolvencies-june-2026"

BPE_XLSX_URL = "https://assets.publishing.service.gov.uk/media/68dbccc9c487360cc70c9f4e/BPE_2025_detailed_tables.xlsx"
BPE_RELEASE_PAGE = "https://www.gov.uk/government/statistics/business-population-estimates-2025"

ONS_SURVIVAL_DATASET_PAGE = (
    "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/"
    "datasets/businessdemographyreferencetable/current"
)

OUT_PATH = (
    Path(__file__).resolve().parents[2]
    / "generalist" / "web" / "src" / "data" / "uk-small-business-barometer.json"
)


def fetch_text(url: str) -> str:
    r = requests.get(url, timeout=60, headers=UA)
    r.raise_for_status()
    return r.text


def fetch_workbook(url: str) -> openpyxl.Workbook:
    r = requests.get(url, timeout=90, headers=UA)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)


# ---------------------------------------------------------------------------
# 1. Companies House quarterly register series (UK, All companies)
# ---------------------------------------------------------------------------

def parse_ch_quarterly() -> tuple[list[dict], dict]:
    text = fetch_text(CH_CSV_URL)
    by_date: dict[str, dict[str, int]] = defaultdict(dict)
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        if row["Region"] == "UK" and row["Corporate body type"] == "All companies":
            by_date[row["Date"]][row["Attribute"]] = int(row["Value"])

    def sort_key(d: str):
        dd, mm, yyyy = d.split("/")
        return (int(yyyy), int(mm), int(dd))

    dates = sorted(by_date, key=sort_key)
    quarterly = []
    for d in dates:
        dd, mm, yyyy = d.split("/")
        vals = by_date[d]
        quarterly.append({
            "quarter_end": f"{yyyy}-{mm}-{dd}",
            "incorporations": vals.get("Incorporations", 0),
            "dissolutions": vals.get("Dissolutions", 0),
            "net": vals.get("Incorporations", 0) - vals.get("Dissolutions", 0),
            "register_total": vals.get("Total register size at end of period", 0),
            "register_effective": vals.get("Effective register size at end of period", 0),
        })

    latest = quarterly[-1]
    # FYE = latest quarter + 3 preceding quarters
    fye_now = quarterly[-4:]
    fye_prior = quarterly[-8:-4]
    inc_now = sum(q["incorporations"] for q in fye_now)
    diss_now = sum(q["dissolutions"] for q in fye_now)
    inc_prior = sum(q["incorporations"] for q in fye_prior)
    diss_prior = sum(q["dissolutions"] for q in fye_prior)

    prior_year_quarter = quarterly[-5]  # same quarter, one year earlier
    register_yoy_pct = round(
        (latest["register_total"] - prior_year_quarter["register_total"])
        / prior_year_quarter["register_total"] * 100, 2
    ) if prior_year_quarter["register_total"] else None

    # Seasonality: average incorporations by calendar quarter across full years
    by_calendar_q: dict[str, list[int]] = defaultdict(list)
    for q in quarterly:
        _, mm, _ = q["quarter_end"].split("-")[0], q["quarter_end"].split("-")[1], q["quarter_end"].split("-")[2]
        month = int(q["quarter_end"].split("-")[1])
        cal_q = {3: "Q1 (Jan-Mar)", 6: "Q2 (Apr-Jun)", 9: "Q3 (Jul-Sep)", 12: "Q4 (Oct-Dec)"}[month]
        by_calendar_q[cal_q].append(q["incorporations"])
    seasonality = [
        {"calendar_quarter": k, "avg_incorporations": round(sum(v) / len(v))}
        for k, v in by_calendar_q.items()
    ]

    headline = {
        "fye_label": f"FYE {latest['quarter_end'][:7]}",
        "incorporations_fye": inc_now,
        "dissolutions_fye": diss_now,
        "incorporations_yoy_pct": round((inc_now - inc_prior) / inc_prior * 100, 2) if inc_prior else None,
        "dissolutions_yoy_pct": round((diss_now - diss_prior) / diss_prior * 100, 2) if diss_prior else None,
        "register_total": latest["register_total"],
        "register_effective": latest["register_effective"],
        "register_yoy_pct": register_yoy_pct,
        "as_of": latest["quarter_end"],
    }
    return quarterly, {"headline": headline, "seasonality": seasonality}


# ---------------------------------------------------------------------------
# 2. Insolvency Service -- all-sector monthly/annual totals (Table A1a, "Total")
# ---------------------------------------------------------------------------

def parse_insolvency_totals(wb: openpyxl.Workbook) -> tuple[list[dict], list[dict], dict]:
    ws = wb["Table_A1a"]
    rows = list(ws.iter_rows(values_only=True))
    header = next(r for r in rows if r[0] == "Section")
    total_row = next(r for r in rows if r[0] == "Total")

    year_cols = [(i, int(h)) for i, h in enumerate(header) if isinstance(h, str) and h.strip().isdigit()]
    month_cols = [(i, h) for i, h in enumerate(header) if isinstance(h, str) and re.match(r"^[A-Z][a-z]{2} \d{4}$", h)]

    annual = [{"year": y, "total": int(total_row[i])} for i, y in year_cols if total_row[i] is not None]

    month_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
        "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }

    def to_ym(label: str) -> str:
        mon, yr = label.split(" ")
        return f"{yr}-{month_map[mon]}"

    monthly = [
        {"month": to_ym(h), "total": int(total_row[i])}
        for i, h in month_cols if total_row[i] is not None
    ]
    monthly.sort(key=lambda r: r["month"])

    last = monthly[-1]
    ly_month = str(int(last["month"][:4]) - 1) + last["month"][4:]
    prev = next((m for m in monthly if m["month"] == ly_month), None)
    yoy = round((last["total"] - prev["total"]) / prev["total"] * 100, 1) if prev and prev["total"] else None
    ttm_total = sum(m["total"] for m in monthly[-12:])

    headline = {
        "last_settled_month": last["month"],
        "last_month_total": last["total"],
        "yoy_pct": yoy,
        "ttm_total": ttm_total,
        "coverage": "England and Wales",
    }
    return monthly, annual, headline


def parse_insolvency_procedure_mix(wb: openpyxl.Workbook) -> dict:
    """All-sector procedure mix (Table 1b): latest month + TTM CVL share."""
    ws = wb["Table_1b"]
    rows = list(ws.iter_rows(values_only=True))
    header = next(r for r in rows if r[0] == "Period")
    col = {name: i for i, name in enumerate(header)}

    month_rows = [
        r for r in rows
        if isinstance(r[0], str) and re.match(r"^[A-Z][a-z]{2} \d{4}$", r[0])
    ]

    def month_key(label: str) -> str:
        mon, yr = label.split(" ")
        months = {
            "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
            "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
        }
        return f"{yr}-{months[mon]}"

    month_rows.sort(key=lambda r: month_key(r[0]))
    last = month_rows[-1]
    ttm_rows = month_rows[-12:]

    def safe(r, name):
        v = r[col[name]]
        return int(v) if isinstance(v, (int, float)) else 0

    ttm_total = sum(safe(r, "Total Company Insolvencies") for r in ttm_rows)
    ttm_cvl = sum(safe(r, "Creditors' voluntary liquidations") for r in ttm_rows)

    return {
        "last_month": month_key(last[0]),
        "last_month_total": safe(last, "Total Company Insolvencies"),
        "last_month_cvl": safe(last, "Creditors' voluntary liquidations"),
        "last_month_cvl_pct": round(safe(last, "Creditors' voluntary liquidations") / safe(last, "Total Company Insolvencies") * 100, 1) if safe(last, "Total Company Insolvencies") else None,
        "ttm_cvl_pct": round(ttm_cvl / ttm_total * 100, 1) if ttm_total else None,
    }


# ---------------------------------------------------------------------------
# 3. ONS Business Demography survival -- Table 4.2, "Total" row (all industries)
# ---------------------------------------------------------------------------

SURVIVAL_COLS = [("y1", 2, 3), ("y2", 4, 5), ("y3", 6, 7), ("y4", 8, 9), ("y5", 10, 11)]


def discover_survival_xlsx_url() -> str:
    r = requests.get(ONS_SURVIVAL_DATASET_PAGE, timeout=30, headers=UA)
    r.raise_for_status()
    matches = re.findall(r'href="(/file\?uri=[^"]+businessdemographyexceltables[^"]+\.xlsx)"', r.text)
    current = [m for m in matches if "/previous/" not in m]
    if not current:
        raise RuntimeError("Could not find current ONS survival XLSX link")
    return "https://www.ons.gov.uk" + current[0]


def parse_survival_total(wb: openpyxl.Workbook) -> tuple[list[dict], dict]:
    ws = wb["Table 4.2"]
    cohorts: list[dict[str, Any]] = []
    cur_year: int | None = None
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        if isinstance(row[0], int) and 2000 < row[0] < 2100:
            cur_year = row[0]
            continue
        if str(row[0]).strip() != "Total" or cur_year is None:
            continue
        births = row[1]
        entry: dict[str, Any] = {"birth_year": cur_year, "births": int(births) if births is not None else None}
        for key, count_off, pct_off in SURVIVAL_COLS:
            pct_v = row[pct_off]
            entry[f"{key}_pct"] = round(float(pct_v), 1) if isinstance(pct_v, (int, float)) else None
        cohorts.append(entry)

    five_yr = [c for c in cohorts if c["y5_pct"] is not None]
    one_yr = [c for c in cohorts if c["y1_pct"] is not None]
    headline = {
        "latest_5yr_cohort_year": five_yr[-1]["birth_year"] if five_yr else None,
        "latest_5yr_pct": five_yr[-1]["y5_pct"] if five_yr else None,
        "latest_1yr_cohort_year": one_yr[-1]["birth_year"] if one_yr else None,
        "latest_1yr_pct": one_yr[-1]["y1_pct"] if one_yr else None,
    }
    return cohorts, headline


# ---------------------------------------------------------------------------
# 4. ONS/DBT Business Population Estimates -- Tables 1, 3, 25
# ---------------------------------------------------------------------------

def parse_bpe(wb: openpyxl.Workbook) -> dict:
    t1 = list(wb["Table 1"].iter_rows(values_only=True))
    all_biz = next(r for r in t1 if r[0] == "All businesses")
    total = int(all_biz[1])
    large = sum(
        int(r[1]) for r in t1
        if r[0] in ("250 to 499", "500 to 999", "1000 or more")
    )
    sme_pct = round((total - large) / total * 100, 2)
    no_employees_rows = [r for r in t1 if isinstance(r[0], str) and "no employees" in r[0]]
    no_employees = sum(int(r[1]) for r in no_employees_rows)
    no_employees_pct = round(no_employees / total * 100, 1)

    t3 = list(wb["Table 3"].iter_rows(values_only=True))

    def legal_status_count(label: str) -> int:
        idx = next(i for i, r in enumerate(t3) if r[0] == label)
        return int(t3[idx + 1][1])  # "All businesses" row directly under the section header

    companies = legal_status_count("Companies (including public corporations and nationalised bodies)[Note 3] [Note 4]")
    partnerships = legal_status_count("Ordinary Partnerships")
    sole_props = legal_status_count("Sole proprietorships")

    t25 = list(wb["Table 25"].iter_rows(values_only=True))
    header = next(r for r in t25 if r[0] == "Employee size band")
    all_biz_ts = next(r for r in t25 if r[0] == "All businesses")
    year_idx = [
        (i, int(h)) for i, h in enumerate(header)
        if isinstance(h, (int, str)) and str(h).strip().isdigit() and len(str(h).strip()) == 4
    ]
    timeseries = [
        {"year": y, "total": int(all_biz_ts[i])}
        for i, y in year_idx if all_biz_ts[i] is not None
    ]
    latest_yr = timeseries[-1]
    prior_yr = timeseries[-2]
    yoy_pct = round((latest_yr["total"] - prior_yr["total"]) / prior_yr["total"] * 100, 2)
    yoy_abs = latest_yr["total"] - prior_yr["total"]

    headline = {
        "total_businesses": total,
        "as_of_start_year": latest_yr["year"],
        "yoy_pct": yoy_pct,
        "yoy_abs": yoy_abs,
        "sme_pct": sme_pct,
        "no_employees_pct": no_employees_pct,
        "companies_count": companies,
        "companies_pct": round(companies / total * 100, 1),
        "sole_proprietorships_count": sole_props,
        "sole_proprietorships_pct": round(sole_props / total * 100, 1),
        "partnerships_count": partnerships,
        "partnerships_pct": round(partnerships / total * 100, 1),
    }
    return {"headline": headline, "timeseries": timeseries}


def main() -> None:
    print("Fetching Companies House quarterly register series...")
    quarterly, ch_bundle = parse_ch_quarterly()
    print(f"  {len(quarterly)} quarters, latest {quarterly[-1]['quarter_end']}")

    print("Fetching Insolvency Service Industry Tables (all-sector totals)...")
    insolvency_wb = fetch_workbook(INSOLVENCY_INDUSTRY_XLSX_URL)
    insolvency_monthly, insolvency_annual, insolvency_headline = parse_insolvency_totals(insolvency_wb)
    print(f"  latest month {insolvency_headline['last_settled_month']}, total {insolvency_headline['last_month_total']}")

    print("Fetching Insolvency Service Data Tables (procedure mix, all-sector)...")
    insolvency_data_wb = fetch_workbook(INSOLVENCY_DATA_XLSX_URL)
    procedure_mix = parse_insolvency_procedure_mix(insolvency_data_wb)
    insolvency_headline["cvl_pct_last_month"] = procedure_mix["last_month_cvl_pct"]
    insolvency_headline["cvl_pct_ttm"] = procedure_mix["ttm_cvl_pct"]
    print(f"  CVL share of last month: {procedure_mix['last_month_cvl_pct']}%, TTM: {procedure_mix['ttm_cvl_pct']}%")

    print("Discovering + fetching ONS Business Demography survival table...")
    survival_url = discover_survival_xlsx_url()
    survival_wb = fetch_workbook(survival_url)
    survival_cohorts, survival_headline = parse_survival_total(survival_wb)
    print(f"  5yr cohort {survival_headline['latest_5yr_cohort_year']}: {survival_headline['latest_5yr_pct']}%")

    print("Fetching ONS/DBT Business Population Estimates...")
    bpe_wb = fetch_workbook(BPE_XLSX_URL)
    bpe = parse_bpe(bpe_wb)
    print(f"  {bpe['headline']['total_businesses']:,} businesses, SME {bpe['headline']['sme_pct']}%")

    snapshot = {
        "meta": {
            "generated_at": date.today().isoformat(),
            "sources": [
                {
                    "name": "Incorporated companies in the UK (quarterly)",
                    "publisher": "Companies House",
                    "url": CH_CSV_URL,
                    "release_page": CH_RELEASE_PAGE,
                    "licence": LICENCE,
                    "retrieved": date.today().isoformat(),
                },
                {
                    "name": "Company Insolvency Statistics -- Industry Tables (Table A1a)",
                    "publisher": "The Insolvency Service",
                    "url": INSOLVENCY_INDUSTRY_XLSX_URL,
                    "release_page": INSOLVENCY_RELEASE_PAGE,
                    "licence": LICENCE,
                    "retrieved": date.today().isoformat(),
                    "coverage": "England and Wales (Scotland and Northern Ireland published separately)",
                },
                {
                    "name": "Company Insolvency Statistics -- Data Tables (Table 1b, procedure mix)",
                    "publisher": "The Insolvency Service",
                    "url": INSOLVENCY_DATA_XLSX_URL,
                    "release_page": INSOLVENCY_RELEASE_PAGE,
                    "licence": LICENCE,
                    "retrieved": date.today().isoformat(),
                    "coverage": "England and Wales (Scotland and Northern Ireland published separately)",
                },
                {
                    "name": "Business Demography, UK -- Table 4.2 (survival of newly born enterprises)",
                    "publisher": "Office for National Statistics",
                    "url": survival_url,
                    "release_page": ONS_SURVIVAL_DATASET_PAGE,
                    "licence": LICENCE,
                    "retrieved": date.today().isoformat(),
                },
                {
                    "name": "Business Population Estimates for the UK and regions",
                    "publisher": "Department for Business and Trade / ONS",
                    "url": BPE_XLSX_URL,
                    "release_page": BPE_RELEASE_PAGE,
                    "licence": LICENCE,
                    "retrieved": date.today().isoformat(),
                },
            ],
            "notes": (
                "This barometer fuses four national, all-sector series and is deliberately "
                "distinct from single-SIC research indexes elsewhere in the estate. Company "
                "insolvency figures cover England and Wales only; Scotland and Northern "
                "Ireland are published separately by the Insolvency Service and are not "
                "included here. Company register and formation figures (Companies House) "
                "cover the whole UK. Business population and survival figures (ONS/DBT) "
                "cover the whole UK. Counts are not rates unless stated; an increase in "
                "insolvency numbers can partly reflect growth in the underlying business "
                "population."
            ),
            "attribution": ATTRIBUTION,
        },
        "headline": {
            "incorporations": ch_bundle["headline"],
            "insolvency": insolvency_headline,
            "survival": survival_headline,
            "population": bpe["headline"],
        },
        "register": {
            "quarterly": quarterly,
            "seasonality": ch_bundle["seasonality"],
        },
        "insolvency": {
            "monthly": insolvency_monthly,
            "annual": insolvency_annual,
        },
        "survival": {
            "cohorts": survival_cohorts,
        },
        "population": {
            "timeseries": bpe["timeseries"],
        },
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n[snapshot] wrote {OUT_PATH}")

    # Self-check
    assert len(quarterly) > 20, "CH quarterly series too short"
    assert all(q["net"] == q["incorporations"] - q["dissolutions"] for q in quarterly), "net mismatch"
    assert len(insolvency_monthly) > 100, "insolvency monthly series too short"
    assert len(insolvency_annual) >= 8, "insolvency annual series too short"
    assert len(survival_cohorts) >= 3, "too few survival cohorts"
    assert bpe["headline"]["companies_pct"] + bpe["headline"]["sole_proprietorships_pct"] + bpe["headline"]["partnerships_pct"] > 99, "legal status shares don't sum close to 100"
    print("[self-check] PASS")
    print(f"Incorporations {ch_bundle['headline']['fye_label']}: {ch_bundle['headline']['incorporations_fye']:,} vs dissolutions {ch_bundle['headline']['dissolutions_fye']:,}")
    print(f"Insolvency TTM total: {insolvency_headline['ttm_total']:,}")
    print(f"Business population: {bpe['headline']['total_businesses']:,} ({bpe['headline']['sme_pct']}% SME)")


if __name__ == "__main__":
    main()
