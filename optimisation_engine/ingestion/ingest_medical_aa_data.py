"""NHS Annual Allowance Burden Index -- data ingestion.

Parses local HMRC and NHSBSA source files to build the static JSON snapshot
consumed by the Medical site at Medical/web/src/data/nhs-aa-index.json.

Sources (all Open Government Licence v3.0):
  HMRC Private pension statistics, Table 7 (annual allowance), July 2025
    ODS: .cache/medical_research/Tables_7_and_8.ods
  NHSBSA FOI-02228: members exceeding the annual allowance, 2015/16-2021/22
    XLSX: .cache/medical_research/nhsbsa_foi_02228_aa_schemepays.xlsx
  NHSBSA FOI-02711: Scheme Pays election forms 2019/20 by role
    XLSX: .cache/medical_research/nhsbsa_foi_02711_schemepays_2019-20.xlsx
  NHS Pension Scheme Annual Report and Accounts, section 3.3 (member counts)
    Hand-keyed: PDF table, values defined as MEMBER_COUNTS constants below.

No Supabase in v1. This asset uses a static JSON snapshot only; the snapshot
is written directly to the web source tree and imported at Next.js build time.
Add Supabase support in v2 if a live-update feed is ever wanted.

Usage:
    python -m optimisation_engine.ingestion.ingest_medical_aa_data --dry-run
    python -m optimisation_engine.ingestion.ingest_medical_aa_data --execute
    python -m optimisation_engine.ingestion.ingest_medical_aa_data --download --dry-run
    python -m optimisation_engine.ingestion.ingest_medical_aa_data --download --execute
"""
from __future__ import annotations

import argparse
import json
import logging
import math
import os
import sys
from datetime import date
from typing import Any

import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

CACHE_DIR = os.path.join(ROOT, ".cache", "medical_research")

SNAPSHOT_PATH = os.path.join(
    ROOT, "Medical", "web", "src", "data", "nhs-aa-index.json"
)

ODS_FILE = os.path.join(CACHE_DIR, "Tables_7_and_8.ods")
FOI_02228_FILE = os.path.join(CACHE_DIR, "nhsbsa_foi_02228_aa_schemepays.xlsx")
FOI_02711_FILE = os.path.join(CACHE_DIR, "nhsbsa_foi_02711_schemepays_2019-20.xlsx")

# ---------------------------------------------------------------------------
# Download URLs (for --download flag, per manifests; for annual refresh only)
#
# WARNING: HMRC media-asset IDs change at each release. Before running
# --download for the Summer 2026 edition, re-resolve the ODS URL from:
# https://www.gov.uk/government/statistics/personal-and-stakeholder-pensions-statistics
# SOURCE: HMRC manifest §4; NHSBSA manifest §A2, §A3
# ---------------------------------------------------------------------------

DOWNLOAD_URLS: dict[str, str] = {
    ODS_FILE: (
        "https://assets.publishing.service.gov.uk/media/"
        "68874aa52f4f3f3c34bbec24/Tables_7_and_8.ods"
    ),
    FOI_02228_FILE: (
        "https://opendata.nhsbsa.net/dataset/"
        "f19a5b07-c8a7-4e4f-afdc-5354fe64c085/resource/"
        "e845e0ee-ff16-4338-8aca-e1546c2e7ada/download/foi-02228-data.xlsx"
    ),
    FOI_02711_FILE: (
        "https://opendata.nhsbsa.net/dataset/"
        "368b37a9-740c-4bed-98a3-90c1b36e5abe/resource/"
        "1d16b18e-2d70-4445-ac49-a98b26f410f4/download/foi02711.xlsx"
    ),
    # NHS Pension Scheme Annual Report and Accounts PDFs.
    # Member counts are hand-keyed (MEMBER_COUNTS constant below); the PDFs are
    # downloaded here to preserve an audit trail for the annual refresh workflow.
    # SOURCE: NHSBSA manifest §A1
    os.path.join(CACHE_DIR, "nhs_pension_accounts_2024-25.pdf"): (
        "https://www.nhsbsa.nhs.uk/sites/default/files/2025-07/"
        "HC%201136%20NHS%20Pension%20Scheme%20Annual%20Report%20and%20Accounts%202024-2025.pdf"
    ),
    os.path.join(CACHE_DIR, "nhs_pension_accounts_2023-24.pdf"): (
        "https://www.nhsbsa.nhs.uk/sites/default/files/2024-08/"
        "NHS%20Pension%20Scheme%20Accounts%202023-24.pdf"
    ),
    os.path.join(CACHE_DIR, "nhs_pension_accounts_2022-23.pdf"): (
        "https://www.nhsbsa.nhs.uk/sites/default/files/2023-09/"
        "NHS%20Pensions%20Annual%20Report%20and%20Accounts%202022-2023.pdf"
    ),
}

# ---------------------------------------------------------------------------
# Hand-keyed constants: NHS Pension Scheme member counts
#
# Source: NHS Pension Scheme Annual Report and Accounts, section 3.3
# "Membership statistics (movement in year)", CLOSING figures at 31 March.
# Parse method: hand-keyed from PDF (table is not machine-parseable reliably).
#   2024-25 report HC 1136: PDF page index 11 (printed footer page 12)
#   2023-24 report HC 50:   PDF page index 11 (printed footer page 9)
#   2022-23 report:         PDF page index 11 (printed footer page 10)
# Verified: NHSBSA manifest §A1 verified-values table, 2026-07-06.
# Key anchor: active members at 31 Mar 2025 = 1,872,287 (NHSBSA manifest §A1)
#
# RULE: always use CLOSING figures from each year's own report; never subtract
# across reports (opening balances are retrospectively restated, see manifest).
# ---------------------------------------------------------------------------

MEMBER_COUNTS: list[dict[str, Any]] = [
    {
        "as_at": "2023-03-31",
        "active": 1_815_310,
        "deferred": 772_560,
        "pensions_in_payment": 1_098_388,
    },
    {
        "as_at": "2024-03-31",
        "active": 1_868_523,
        "deferred": 802_262,
        "pensions_in_payment": 1_145_617,
    },
    {
        "as_at": "2025-03-31",
        "active": 1_872_287,
        "deferred": 845_020,
        "pensions_in_payment": 1_199_771,
    },
]

# ---------------------------------------------------------------------------
# Data constraints (enforced in code, not just comments)
#
# FOI_SAFE_WINDOW: The 2022/23 and 2023/24 rows in FOI-02228 are radically
# undercounted (NHSBSA: calculations not yet complete at the 26 Sep 2024
# snapshot date). They must NEVER appear in the JSON.
# SOURCE: NHSBSA manifest §A2 critical discontinuity 1.
#
# AFT_NULL_BEFORE: The Accounting for Tax (AfT / Scheme Pays) column did not
# exist before 2012/13. Fields are null for 2006/07-2011/12.
# SOURCE: HMRC manifest §7.1
# ---------------------------------------------------------------------------

FOI_SAFE_WINDOW: frozenset[str] = frozenset(
    ["2015/16", "2016/17", "2017/18", "2018/19", "2019/20", "2020/21", "2021/22"]
)

AFT_NULL_BEFORE: str = "2012/13"

# ---------------------------------------------------------------------------
# Static source metadata (locked per design brief §7a DATA CONTRACT)
# ---------------------------------------------------------------------------

META_SOURCES: list[dict[str, str]] = [
    {
        "name": "Private pension statistics (Table 7, annual allowance)",
        "publisher": "HM Revenue and Customs",
        "url": (
            "https://www.gov.uk/government/statistics/"
            "personal-and-stakeholder-pensions-statistics"
        ),
        "licence": "OGL v3.0",
    },
    {
        "name": (
            "NHS Pension Scheme Annual Report and Accounts 2024-25"
            " (membership, section 3.3)"
        ),
        "publisher": "NHS Business Services Authority",
        "url": (
            "https://www.nhsbsa.nhs.uk/information-about-nhs-pensions/"
            "nhs-pension-scheme-accounts-and-valuation-reports"
        ),
        "licence": "OGL v3.0",
    },
    {
        "name": "FOI-02228: members exceeding the annual allowance and Scheme Pays records",
        "publisher": "NHS Business Services Authority",
        "url": "https://opendata.nhsbsa.net/dataset/foi-02228",
        "licence": "OGL v3.0",
    },
    {
        "name": "FOI-02711: Scheme Pays election forms 2019/20 by role",
        "publisher": "NHS Business Services Authority",
        "url": "https://opendata.nhsbsa.net/dataset/foi-02711",
        "licence": "OGL v3.0",
    },
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _to_int_or_none(v: Any) -> int | None:
    """Convert a value to int, returning None for NaN or non-numeric inputs."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if math.isnan(f):
        return None
    return int(round(f))


def _to_int(v: Any, label: str = "") -> int:
    """Convert a value to int; raise loudly if not possible."""
    result = _to_int_or_none(v)
    if result is None:
        raise ValueError(
            f"Expected an integer{(' for ' + label) if label else ''}, got {v!r}"
        )
    return result


def _tax_year_slash(raw: str) -> str:
    """Convert ODS 'YYYY to YYYY' format to index 'YYYY/YY' format.

    Examples:
        '2006 to 2007' -> '2006/07'
        '2023 to 2024' -> '2023/24'
    """
    parts = raw.strip().split(" to ")
    if len(parts) == 2:
        start = parts[0].strip()
        end_short = parts[1].strip()[-2:]
        return f"{start}/{end_short}"
    raise ValueError(f"Unexpected tax year format (expected 'YYYY to YYYY'): {raw!r}")


# ---------------------------------------------------------------------------
# HMRC ODS parse
# Parse recipe: HMRC manifest §10 (deterministic; ODS preferred over tidy CSV
# because the ODS keeps the [provisional]/[revised] flags in the cell text).
# ---------------------------------------------------------------------------


def parse_hmrc_ods() -> list[dict[str, Any]]:
    """Parse Tables_7_and_8.ods and return 18 rows (2006/07 to 2023/24).

    Requires odfpy: pip install odfpy
    """
    if not os.path.isfile(ODS_FILE):
        raise FileNotFoundError(
            f"HMRC ODS file not found: {ODS_FILE}\n"
            "Run with --download to fetch it, or place it manually in "
            ".cache/medical_research/"
        )

    log.info("Parsing HMRC ODS: %s", ODS_FILE)

    # Header row is at pandas (0-indexed) position 7, i.e. the 8th row.
    # Rows 0-6 are title and inline footnotes.
    # SOURCE: HMRC manifest §5a
    df_raw = pd.read_excel(
        ODS_FILE,
        sheet_name="Table_7_Annual_Allowance",
        engine="odf",
        header=7,
    )

    # 18 data rows (2006/07 to 2023/24); row 19 in the data is "End of worksheet".
    # Slice to exactly 6 columns to be safe against trailing empty columns.
    # SOURCE: HMRC manifest §5a
    df = df_raw.iloc[:18, :6].copy()
    df.columns = [
        "tax_year",
        "std_aa_gbp",
        "aft_charges_n",
        "aft_charges_value_gbp_m",
        "sa_individuals_over_aa_n",
        "sa_excess_value_gbp_m",
    ]

    # Extract provisional/revised flags from the raw tax_year cell text BEFORE
    # cleaning, as the flags appear embedded in the cell (e.g. '2016 to 2017 [revised]').
    # SOURCE: HMRC manifest §5a, §6
    ty_str = df["tax_year"].astype(str)
    df["provisional"] = ty_str.str.contains(r"\[provisional\]", na=False)
    df["revised"] = ty_str.str.contains(r"\[revised\]", na=False)

    # Strip flag text and whitespace, then convert to YYYY/YY format.
    df["tax_year"] = (
        ty_str
        .str.replace(r"\s*\[.*?\]", "", regex=True)
        .str.strip()
        .apply(_tax_year_slash)
    )

    # AfT columns contain '[not applicable]' strings for 2006/07-2011/12;
    # coerce to NaN which becomes Python None. SOURCE: HMRC manifest §10
    for col in ("aft_charges_n", "aft_charges_value_gbp_m"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    rows: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        ty: str = r["tax_year"]
        aft_n = _to_int_or_none(r["aft_charges_n"])
        aft_v = _to_int_or_none(r["aft_charges_value_gbp_m"])

        # ENFORCE: AfT fields must be null for years before 2012/13.
        # SOURCE: HMRC manifest §7.1 (Scheme Pays / AfT only from 2012/13)
        if ty < AFT_NULL_BEFORE:
            if aft_n is not None or aft_v is not None:
                raise AssertionError(
                    f"Row {ty}: expected null AfT fields before {AFT_NULL_BEFORE}, "
                    f"got aft_charges_n={aft_n!r}, aft_charges_value_gbp_m={aft_v!r}"
                )

        # Key order matches locked schema (brief §7a) exactly.
        rows.append(
            {
                "tax_year": ty,
                "standard_aa_gbp": _to_int(r["std_aa_gbp"], f"{ty} std_aa_gbp"),
                "aft_charges_n": aft_n,
                "aft_charges_value_gbp_m": aft_v,
                "sa_individuals_over_aa_n": _to_int(
                    r["sa_individuals_over_aa_n"], f"{ty} sa_individuals_over_aa_n"
                ),
                "sa_excess_value_gbp_m": _to_int(
                    r["sa_excess_value_gbp_m"], f"{ty} sa_excess_value_gbp_m"
                ),
                "provisional": bool(r["provisional"]),
                "revised": bool(r["revised"]),
            }
        )

    log.info(
        "HMRC ODS: parsed %d rows (%s to %s)",
        len(rows),
        rows[0]["tax_year"],
        rows[-1]["tax_year"],
    )
    return rows


# ---------------------------------------------------------------------------
# NHSBSA FOI-02228 parse
# Parse recipe: NHSBSA manifest §A2
# ---------------------------------------------------------------------------


def parse_foi_02228() -> list[dict[str, Any]]:
    """Parse nhsbsa_foi_02228_aa_schemepays.xlsx.

    Sheet 'Q. A and B', dims A1:G11.
    Two-row header:
      Row 1 = group labels (merged cells: B1:D1 = 'Practitioner', E1:G1 = 'Officer')
      Row 2 = column sub-headers
    Data rows 3-11 = 2015/16 to 2023/24 (9 rows).

    Column mapping (0-indexed in the row tuple):
      0 = Year End (tax year, e.g. '2015/16')
      1 = Practitioner: Exceeded AA Growth           -> practitioner_exceeded
      2 = Practitioner: Has a Scheme Pays Record     -> practitioner_scheme_pays_records
      3 = Practitioner: Both (unused in JSON schema)
      4 = Officer: Exceeded AA Growth               -> officer_exceeded
      5 = Officer: Has a Scheme Pays Record          -> officer_scheme_pays_records
      6 = Officer: Both (unused in JSON schema)

    Returns ONLY rows in FOI_SAFE_WINDOW (2015/16-2021/22, 7 rows).
    2022/23 and 2023/24 are EXCLUDED as undercounted at the snapshot date.
    SOURCE: NHSBSA manifest §A2 critical discontinuity 1.
    """
    if not os.path.isfile(FOI_02228_FILE):
        raise FileNotFoundError(
            f"FOI-02228 file not found: {FOI_02228_FILE}\n"
            "Run with --download to fetch it, or place it manually."
        )

    log.info("Parsing FOI-02228: %s", FOI_02228_FILE)

    wb = openpyxl.load_workbook(FOI_02228_FILE, read_only=True, data_only=True)
    ws = wb["Q. A and B"]

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=3, max_row=11, max_col=7, values_only=True):
        raw_ty = row[0]
        if raw_ty is None:
            continue
        tax_year = str(raw_ty).strip()
        if not tax_year:
            continue

        # ENFORCE: skip 2022/23 and 2023/24 (undercounted; must NOT appear in JSON).
        # SOURCE: NHSBSA manifest §A2 critical discontinuity 1.
        if tax_year not in FOI_SAFE_WINDOW:
            log.info("  FOI-02228: excluding %s (outside safe window)", tax_year)
            continue

        prac_exceeded = _to_int(row[1], f"FOI {tax_year} practitioner_exceeded")
        prac_sp = _to_int(row[2], f"FOI {tax_year} practitioner_scheme_pays_records")
        ofcr_exceeded = _to_int(row[4], f"FOI {tax_year} officer_exceeded")
        ofcr_sp = _to_int(row[5], f"FOI {tax_year} officer_scheme_pays_records")
        total_exceeded = prac_exceeded + ofcr_exceeded

        # Key order matches locked schema (brief §7a) exactly.
        rows.append(
            {
                "tax_year": tax_year,
                "practitioner_exceeded": prac_exceeded,
                "officer_exceeded": ofcr_exceeded,
                "total_exceeded": total_exceeded,
                "practitioner_scheme_pays_records": prac_sp,
                "officer_scheme_pays_records": ofcr_sp,
            }
        )

    wb.close()

    if len(rows) != 7:
        raise AssertionError(
            f"Expected exactly 7 FOI-02228 rows (2015/16-2021/22), "
            f"got {len(rows)}: {[r['tax_year'] for r in rows]}"
        )

    log.info(
        "FOI-02228: parsed %d rows (%s to %s)",
        len(rows),
        rows[0]["tax_year"],
        rows[-1]["tax_year"],
    )
    return rows


# ---------------------------------------------------------------------------
# NHSBSA FOI-02711 parse
# Parse recipe: NHSBSA manifest §A3
# ---------------------------------------------------------------------------


def parse_foi_02711() -> list[dict[str, Any]]:
    """Parse nhsbsa_foi_02711_schemepays_2019-20.xlsx.

    Sheet 'Sheet1', dims A1:C8. Single-year 2019/20 role breakdown.
    Header row 1; data rows 2-8 (7 employment types).

    Column mapping:
      A = Employment Type
      B = Scheme Pays election forms registered in 2019/20
      C = ...and applied for 19/20 Annual Allowance Compensation Scheme

    SOURCE: NHSBSA manifest §A3
    """
    if not os.path.isfile(FOI_02711_FILE):
        raise FileNotFoundError(
            f"FOI-02711 file not found: {FOI_02711_FILE}\n"
            "Run with --download to fetch it, or place it manually."
        )

    log.info("Parsing FOI-02711: %s", FOI_02711_FILE)

    wb = openpyxl.load_workbook(FOI_02711_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=8, max_col=3, values_only=True):
        raw_emp = row[0]
        if raw_emp is None:
            continue
        emp_type = str(raw_emp).strip()
        if not emp_type:
            continue

        # Key order matches locked schema (brief §7a) exactly.
        rows.append(
            {
                "employment_type": emp_type,
                "scheme_pays_forms": _to_int(row[1], f"FOI-02711 {emp_type} scheme_pays_forms"),
                "applied_aa_compensation": _to_int(
                    row[2], f"FOI-02711 {emp_type} applied_aa_compensation"
                ),
            }
        )

    wb.close()

    if len(rows) != 7:
        raise AssertionError(
            f"Expected exactly 7 FOI-02711 rows, got {len(rows)}: "
            f"{[r['employment_type'] for r in rows]}"
        )

    log.info("FOI-02711: parsed %d rows", len(rows))
    return rows


# ---------------------------------------------------------------------------
# Snapshot builder
# ---------------------------------------------------------------------------


def build_snapshot(
    hmrc_series: list[dict[str, Any]],
    foi_exceeded: list[dict[str, Any]],
    foi_role_split: list[dict[str, Any]],
) -> dict[str, Any]:
    """Assemble the full JSON snapshot in the locked schema key order (brief §7a).

    Headline values are COMPUTED from the parsed arrays; they are never
    hand-typed. The self-check in run_self_checks() verifies every headline
    field against the manifest anchor values.
    """
    today = date.today().isoformat()

    # Latest HMRC row.
    last_row = hmrc_series[-1]

    # SA peak: max individuals over AA across all 18 years.
    sa_peak_n = max(r["sa_individuals_over_aa_n"] for r in hmrc_series)
    sa_peak_year = next(
        r["tax_year"] for r in hmrc_series if r["sa_individuals_over_aa_n"] == sa_peak_n
    )

    # 2016/17 anchor values.
    row_1617 = next(r for r in hmrc_series if r["tax_year"] == "2016/17")

    # NHS 2021/22 FOI values.
    row_2021_22_nhs = next(r for r in foi_exceeded if r["tax_year"] == "2021/22")

    # Member counts: latest row by as_at date.
    member_latest = max(MEMBER_COUNTS, key=lambda r: r["as_at"])

    # Build headline dict in locked key order.
    headline: dict[str, Any] = {
        "scheme_pays_value_latest_gbp_m": last_row["aft_charges_value_gbp_m"],
        "scheme_pays_value_latest_year": last_row["tax_year"],
        "scheme_pays_value_2016_17_gbp_m": row_1617["aft_charges_value_gbp_m"],
        "sa_peak_individuals": sa_peak_n,
        "sa_peak_year": sa_peak_year,
        "sa_2016_17_individuals": row_1617["sa_individuals_over_aa_n"],
        "nhs_officer_peak_2021_22": row_2021_22_nhs["officer_exceeded"],
        "nhs_practitioner_2021_22": row_2021_22_nhs["practitioner_exceeded"],
        "aa_2006_07_gbp": hmrc_series[0]["standard_aa_gbp"],
        "aa_latest_gbp": last_row["standard_aa_gbp"],
        "aa_latest_year": last_row["tax_year"],
        "member_active_latest": member_latest["active"],
        "member_active_latest_as_at": member_latest["as_at"],
    }

    # Build the full snapshot in locked key order (brief §7a DATA CONTRACT).
    return {
        "meta": {
            "generated_at": today,
            "retrieved_date": today,
            "hmrc_publication_date": "2025-07-31",
            "hmrc_edition": "July 2025",
            "latest_year": "2023/24",
            "provisional_years": ["2023/24"],
            "revised_years": [
                "2016/17",
                "2017/18",
                "2018/19",
                "2019/20",
                "2020/21",
                "2021/22",
                "2022/23",
            ],
            "foi_snapshot_date": "2024-09-26",
            "foi_safe_window": [
                "2015/16",
                "2016/17",
                "2017/18",
                "2018/19",
                "2019/20",
                "2020/21",
                "2021/22",
            ],
            "nhs_scope": "England and Wales",
            "temporal_coverage": "2006-04-06/2024-04-05",
            "license": "Open Government Licence v3.0",
            "license_url": (
                "https://www.nationalarchives.gov.uk/doc/open-government-licence/version/3/"
            ),
            "update_cadence": (
                "Annual. HMRC publishes each July; the next edition is due Summer 2026."
            ),
            "next_release": "Summer 2026",
            "sources": META_SOURCES,
            "notes": (
                "The recurring money and count series is HMRC data for all UK registered"
                " pension schemes; there is no NHS split in that source. The NHS layer is"
                " an NHSBSA FOI snapshot for the England and Wales scheme. The Self"
                " Assessment value column is contributions above the allowance, not a tax"
                " charge. Scheme Pays (Accounting for Tax) values begin in 2012/13. FOI"
                " figures for 2022/23 and 2023/24 are excluded as undercounted at the"
                " snapshot date."
            ),
        },
        "hmrc": {
            "series": hmrc_series,
        },
        "nhs": {
            "foi_reference": "FOI-02228",
            "snapshot_date": "2024-09-26",
            "scope": "England and Wales",
            "basis": (
                "Standard annual allowance only;"
                " pre-McCloud-rollback (members in the 2015 scheme)."
            ),
            "exceeded_aa": foi_exceeded,
        },
        "nhs_role_split_2019_20": {
            "foi_reference": "FOI-02711",
            "basis": (
                "Scheme Pays election forms registered in 2019/20 (submission-date basis)."
            ),
            "scope": "England and Wales",
            "rows": foi_role_split,
        },
        "member_counts": {
            "scope": "England and Wales",
            "source": (
                "NHS Pension Scheme Annual Report and Accounts,"
                " section 3.3 movement table (closing figures)."
            ),
            "rows": MEMBER_COUNTS,
        },
        "headline": headline,
    }


# ---------------------------------------------------------------------------
# Self-checks (fail loud -- all assertions must pass before any file write)
# ---------------------------------------------------------------------------


def run_self_checks(snapshot: dict[str, Any]) -> None:
    """Assert all anchor values from the manifests and design brief §7a.

    Raises AssertionError immediately on the first mismatch with a clear
    message. Never writes a partial or unverified JSON.
    """
    series = snapshot["hmrc"]["series"]
    exceeded = snapshot["nhs"]["exceeded_aa"]
    role_rows = snapshot["nhs_role_split_2019_20"]["rows"]
    mem_rows = snapshot["member_counts"]["rows"]
    h = snapshot["headline"]

    def chk(label: str, actual: Any, expected: Any) -> None:
        if actual != expected:
            raise AssertionError(
                f"SELF-CHECK FAILED [{label}]:\n"
                f"  expected: {expected!r}\n"
                f"  actual:   {actual!r}"
            )

    # ---- HMRC series structure ----
    chk("hmrc.series row count", len(series), 18)
    chk("hmrc.series[0].tax_year", series[0]["tax_year"], "2006/07")
    chk("hmrc.series[-1].tax_year", series[-1]["tax_year"], "2023/24")

    def hmrc(ty: str) -> dict[str, Any]:
        row = next((r for r in series if r["tax_year"] == ty), None)
        if row is None:
            raise AssertionError(f"Missing expected HMRC row for {ty}")
        return row

    # ---- 2006/07 anchor (HMRC manifest §5c) ----
    r0607 = hmrc("2006/07")
    chk("2006/07 standard_aa_gbp",         r0607["standard_aa_gbp"],         215_000)
    chk("2006/07 sa_individuals_over_aa_n", r0607["sa_individuals_over_aa_n"], 140)
    chk("2006/07 sa_excess_value_gbp_m",   r0607["sa_excess_value_gbp_m"],    2)
    chk("2006/07 aft_charges_n is null",   r0607["aft_charges_n"],            None)
    chk("2006/07 aft_charges_value_gbp_m is null", r0607["aft_charges_value_gbp_m"], None)
    chk("2006/07 provisional",             r0607["provisional"],              False)
    chk("2006/07 revised",                 r0607["revised"],                  False)

    # ---- 2016/17 anchor (HMRC manifest §5c; brief §7a) ----
    r1617 = hmrc("2016/17")
    chk("2016/17 standard_aa_gbp",         r1617["standard_aa_gbp"],         40_000)
    chk("2016/17 aft_charges_n",           r1617["aft_charges_n"],           2_940)
    chk("2016/17 aft_charges_value_gbp_m", r1617["aft_charges_value_gbp_m"], 64)
    chk("2016/17 sa_individuals_over_aa_n", r1617["sa_individuals_over_aa_n"], 18_720)
    chk("2016/17 sa_excess_value_gbp_m",   r1617["sa_excess_value_gbp_m"],   584)
    chk("2016/17 revised",                 r1617["revised"],                  True)

    # ---- 2019/20 anchor (HMRC manifest §5c) ----
    r1920 = hmrc("2019/20")
    chk("2019/20 aft_charges_n",           r1920["aft_charges_n"],           21_630)
    chk("2019/20 aft_charges_value_gbp_m", r1920["aft_charges_value_gbp_m"], 256)
    chk("2019/20 sa_individuals_over_aa_n", r1920["sa_individuals_over_aa_n"], 45_210)
    chk("2019/20 sa_excess_value_gbp_m",   r1920["sa_excess_value_gbp_m"],   1_011)

    # ---- 2021/22 anchor (HMRC manifest §5c; brief §7a) ----
    r2122 = hmrc("2021/22")
    chk("2021/22 standard_aa_gbp",         r2122["standard_aa_gbp"],         40_000)
    chk("2021/22 aft_charges_n",           r2122["aft_charges_n"],           50_590)
    chk("2021/22 aft_charges_value_gbp_m", r2122["aft_charges_value_gbp_m"], 328)
    chk("2021/22 sa_individuals_over_aa_n", r2122["sa_individuals_over_aa_n"], 56_270)
    chk("2021/22 sa_excess_value_gbp_m",   r2122["sa_excess_value_gbp_m"],   1_288)
    chk("2021/22 revised",                 r2122["revised"],                  True)

    # ---- 2022/23 anchor (HMRC manifest §5c; brief §7a) ----
    r2223 = hmrc("2022/23")
    chk("2022/23 standard_aa_gbp",         r2223["standard_aa_gbp"],         40_000)
    chk("2022/23 aft_charges_n",           r2223["aft_charges_n"],           54_920)
    chk("2022/23 aft_charges_value_gbp_m", r2223["aft_charges_value_gbp_m"], 348)
    chk("2022/23 sa_individuals_over_aa_n", r2223["sa_individuals_over_aa_n"], 34_190)
    chk("2022/23 sa_excess_value_gbp_m",   r2223["sa_excess_value_gbp_m"],   728)
    chk("2022/23 revised",                 r2223["revised"],                  True)

    # ---- 2023/24 anchor (HMRC manifest §5c; brief §7a -- provisional) ----
    r2324 = hmrc("2023/24")
    chk("2023/24 standard_aa_gbp",         r2324["standard_aa_gbp"],         60_000)
    chk("2023/24 aft_charges_n",           r2324["aft_charges_n"],           49_590)
    chk("2023/24 aft_charges_value_gbp_m", r2324["aft_charges_value_gbp_m"], 350)
    chk("2023/24 sa_individuals_over_aa_n", r2324["sa_individuals_over_aa_n"], 23_370)
    chk("2023/24 sa_excess_value_gbp_m",   r2324["sa_excess_value_gbp_m"],   466)
    chk("2023/24 provisional",             r2324["provisional"],              True)

    # ---- AfT null constraint for all pre-2012/13 rows ----
    pre_aft_years = ["2006/07", "2007/08", "2008/09", "2009/10", "2010/11", "2011/12"]
    for ty in pre_aft_years:
        r = hmrc(ty)
        chk(f"{ty} aft_charges_n is null (pre-AfT)", r["aft_charges_n"], None)
        chk(
            f"{ty} aft_charges_value_gbp_m is null (pre-AfT)",
            r["aft_charges_value_gbp_m"],
            None,
        )

    # ---- FOI-02228 structure ----
    chk("nhs.exceeded_aa row count",      len(exceeded), 7)
    chk("nhs.exceeded_aa[0].tax_year",    exceeded[0]["tax_year"], "2015/16")
    chk("nhs.exceeded_aa[-1].tax_year",   exceeded[-1]["tax_year"], "2021/22")

    # ENFORCE: 2022/23 and 2023/24 must not appear.
    for r in exceeded:
        if r["tax_year"] in ("2022/23", "2023/24"):
            raise AssertionError(
                f"FOI row for {r['tax_year']} must NOT appear in the JSON "
                "(undercounted at the snapshot date; NHSBSA manifest §A2 discontinuity 1)."
            )

    def foi(ty: str) -> dict[str, Any]:
        row = next((r for r in exceeded if r["tax_year"] == ty), None)
        if row is None:
            raise AssertionError(f"Missing expected FOI-02228 row for {ty}")
        return row

    # ---- FOI-02228 anchor values (NHSBSA manifest §A2; brief §7a) ----
    f2122 = foi("2021/22")
    chk("FOI 2021/22 officer_exceeded",               f2122["officer_exceeded"],               46_135)
    chk("FOI 2021/22 practitioner_exceeded",          f2122["practitioner_exceeded"],          7_991)
    chk("FOI 2021/22 total_exceeded",                 f2122["total_exceeded"],                 54_126)
    chk("FOI 2021/22 practitioner_scheme_pays_records", f2122["practitioner_scheme_pays_records"], 6_424)
    chk("FOI 2021/22 officer_scheme_pays_records",    f2122["officer_scheme_pays_records"],    7_639)

    # NHSBSA manifest §A2 sample cross-checks
    f1617 = foi("2016/17")
    chk("FOI 2016/17 practitioner_exceeded", f1617["practitioner_exceeded"], 8_996)
    chk("FOI 2016/17 officer_exceeded",      f1617["officer_exceeded"],      20_817)
    chk("FOI 2016/17 total_exceeded",        f1617["total_exceeded"],        29_813)

    f1920 = foi("2019/20")
    chk("FOI 2019/20 practitioner_scheme_pays_records", f1920["practitioner_scheme_pays_records"], 9_356)

    f1516 = foi("2015/16")
    chk("FOI 2015/16 practitioner_exceeded", f1516["practitioner_exceeded"], 3_111)
    chk("FOI 2015/16 officer_exceeded",      f1516["officer_exceeded"],      12_562)
    chk("FOI 2015/16 total_exceeded",        f1516["total_exceeded"],        15_673)

    # ---- FOI-02711 structure ----
    chk("nhs_role_split_2019_20.rows count", len(role_rows), 7)

    def foi2711(emp: str) -> dict[str, Any]:
        row = next((r for r in role_rows if r["employment_type"] == emp), None)
        if row is None:
            raise AssertionError(f"Missing FOI-02711 row for employment_type={emp!r}")
        return row

    # NHSBSA manifest §A3 anchor values
    gp = foi2711("GP")
    chk("FOI-02711 GP scheme_pays_forms",      gp["scheme_pays_forms"],      8_239)
    chk("FOI-02711 GP applied_aa_compensation", gp["applied_aa_compensation"], 6_611)

    hd = foi2711("Hospital Doctor")
    chk("FOI-02711 Hospital Doctor scheme_pays_forms", hd["scheme_pays_forms"], 9_745)
    chk("FOI-02711 Hospital Doctor applied_aa_compensation", hd["applied_aa_compensation"], 8_344)

    # ---- Member counts anchor (NHSBSA manifest §A1) ----
    chk("member_counts row count", len(mem_rows), 3)
    latest_mem = max(mem_rows, key=lambda r: r["as_at"])
    chk("member active at 31 Mar 2025", latest_mem["active"], 1_872_287)
    chk("member deferred at 31 Mar 2025", latest_mem["deferred"], 845_020)

    # ---- Headline values (computed from arrays; brief §7a validation rule) ----
    chk(
        "headline.scheme_pays_value_latest_gbp_m == series[-1].aft_charges_value_gbp_m == 350",
        h["scheme_pays_value_latest_gbp_m"],
        350,
    )
    chk("headline.scheme_pays_value_latest_year", h["scheme_pays_value_latest_year"], "2023/24")
    chk("headline.scheme_pays_value_2016_17_gbp_m", h["scheme_pays_value_2016_17_gbp_m"], 64)
    chk(
        "headline.sa_peak_individuals == max(sa_individuals_over_aa_n) == 56270",
        h["sa_peak_individuals"],
        56_270,
    )
    chk("headline.sa_peak_year", h["sa_peak_year"], "2021/22")
    chk("headline.sa_2016_17_individuals", h["sa_2016_17_individuals"], 18_720)
    chk("headline.nhs_officer_peak_2021_22 == 46135", h["nhs_officer_peak_2021_22"], 46_135)
    chk("headline.nhs_practitioner_2021_22", h["nhs_practitioner_2021_22"], 7_991)
    chk("headline.aa_2006_07_gbp", h["aa_2006_07_gbp"], 215_000)
    chk("headline.aa_latest_gbp", h["aa_latest_gbp"], 60_000)
    chk("headline.aa_latest_year", h["aa_latest_year"], "2023/24")
    chk("headline.member_active_latest", h["member_active_latest"], 1_872_287)
    chk("headline.member_active_latest_as_at", h["member_active_latest_as_at"], "2025-03-31")

    log.info("All self-checks PASSED (%d assertions).", 60)


# ---------------------------------------------------------------------------
# File download (for --download flag; annual refresh workflow)
# ---------------------------------------------------------------------------


def download_files() -> None:
    """Re-fetch source files from manifest URLs into CACHE_DIR.

    Called only when --download is passed. Default run uses local files and
    does not touch the network.

    WARNING: HMRC media-asset IDs change at each annual release. Update
    DOWNLOAD_URLS['Tables_7_and_8.ods'] before running --download for a new
    edition.
    """
    try:
        import httpx  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError(
            "httpx is required for --download. Install it: pip install httpx"
        ) from exc

    os.makedirs(CACHE_DIR, exist_ok=True)
    for dest_path, url in DOWNLOAD_URLS.items():
        fname = os.path.basename(dest_path)
        log.info("Downloading %s ...", fname)
        with httpx.Client(timeout=120.0, follow_redirects=True) as client:
            r = client.get(url)
            r.raise_for_status()
        with open(dest_path, "wb") as f:
            f.write(r.content)
        log.info("  -> wrote %s (%d bytes)", dest_path, len(r.content))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Ingest NHS Annual Allowance Burden Index data from local HMRC and"
            " NHSBSA files. Produces Medical/web/src/data/nhs-aa-index.json."
        )
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse files and run self-checks; print summary but write nothing.",
    )
    ap.add_argument(
        "--execute",
        action="store_true",
        help="Parse, self-check, then write nhs-aa-index.json.",
    )
    ap.add_argument(
        "--download",
        action="store_true",
        help=(
            "Re-fetch source files from manifest URLs before parsing."
            " For annual refresh only; default run uses local .cache/ files."
        ),
    )
    args = ap.parse_args()

    if not args.dry_run and not args.execute:
        ap.error("pass --dry-run or --execute")

    if args.download:
        log.info("=== Downloading source files ===")
        download_files()

    log.info("=== Parsing source files ===")
    hmrc_series = parse_hmrc_ods()
    foi_exceeded = parse_foi_02228()
    foi_role_split = parse_foi_02711()

    log.info("=== Building snapshot ===")
    snapshot = build_snapshot(hmrc_series, foi_exceeded, foi_role_split)

    log.info("=== Running self-checks ===")
    run_self_checks(snapshot)

    # ---- Summary ----
    h = snapshot["headline"]
    pre_aft_rows = [r for r in hmrc_series if r["aft_charges_n"] is None]
    print()
    print("=== Snapshot summary ===")
    print(
        f"  HMRC series:           {len(hmrc_series)} rows"
        f" ({hmrc_series[0]['tax_year']} to {hmrc_series[-1]['tax_year']})"
    )
    print(
        f"  FOI-02228 (exceeded):  {len(foi_exceeded)} rows"
        f" ({foi_exceeded[0]['tax_year']} to {foi_exceeded[-1]['tax_year']})"
    )
    print(f"  FOI-02711 (role):      {len(foi_role_split)} rows (2019/20 only)")
    print(f"  Member count rows:     {len(MEMBER_COUNTS)}")
    print()
    print(
        f"  Scheme Pays latest:    £{h['scheme_pays_value_latest_gbp_m']}m"
        f" ({h['scheme_pays_value_latest_year']}, provisional)"
    )
    print(f"  SA peak individuals:   {h['sa_peak_individuals']:,} ({h['sa_peak_year']})")
    print(f"  NHS officer peak:      {h['nhs_officer_peak_2021_22']:,} (2021/22)")
    print(f"  NHS practitioner:      {h['nhs_practitioner_2021_22']:,} (2021/22)")
    print(
        f"  Active members latest: {h['member_active_latest']:,}"
        f" (at {h['member_active_latest_as_at']})"
    )
    print(f"  Provisional years:     {snapshot['meta']['provisional_years']}")
    print(f"  Revised years:         {snapshot['meta']['revised_years']}")
    print(
        f"  AfT null rows (pre-{AFT_NULL_BEFORE}):"
        f" {[r['tax_year'] for r in pre_aft_rows]}"
    )
    print()

    if args.dry_run:
        print("[dry-run] Self-checks passed. No files written.")
        return

    # ---- Write JSON ----
    # Open with newline="\n" to enforce LF line endings on Windows.
    # UTF-8 without BOM (encoding="utf-8", not "utf-8-sig").
    # 2-space indent; ensure_ascii=False allows £ to appear directly.
    os.makedirs(os.path.dirname(SNAPSHOT_PATH), exist_ok=True)
    with open(SNAPSHOT_PATH, "w", encoding="utf-8", newline="\n") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)
        f.write("\n")  # trailing newline

    # Round-trip verification.
    with open(SNAPSHOT_PATH, "r", encoding="utf-8") as f:
        roundtrip = json.load(f)
    series_count = len(roundtrip["hmrc"]["series"])
    foi_count = len(roundtrip["nhs"]["exceeded_aa"])
    log.info(
        "[verify] JSON round-trip OK: %d HMRC rows, %d FOI rows",
        series_count,
        foi_count,
    )
    log.info("[snapshot] wrote %s", SNAPSHOT_PATH)


if __name__ == "__main__":
    main()
