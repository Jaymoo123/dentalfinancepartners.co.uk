"""UK Business Density Map (generalist) -- data ingestion.

Where Britain's small businesses actually are: business count and density
per 10,000 resident adults, by UK region/nation, plus the local employment
size mix (zero-employee share vs 1-49 vs 50-249 vs 250+).

Source:
  Business Population Estimates for the UK and regions, Table 8 (rates per
  10,000 resident adults by region) and Table 9 (regional size summary).
  Department for Business and Trade / Office for National Statistics.
  Licence: Open Government Licence v3.0.

Usage:
  python -m optimisation_engine.ingestion.ingest_generalist_business_density

Self-contained: no Supabase, no git commit.
"""
from __future__ import annotations

import io
import json
from datetime import date
from pathlib import Path

import openpyxl
import requests

UA = {"User-Agent": "Mozilla/5.0"}
LICENCE = "Open Government Licence v3.0"
PUBLISHER = "Department for Business and Trade / Office for National Statistics"
RELEASE_PAGE = "https://www.gov.uk/government/statistics/business-population-estimates-2025"
XLSX_URL = "https://assets.publishing.service.gov.uk/media/68dbccc9c487360cc70c9f4e/BPE_2025_detailed_tables.xlsx"
ATTRIBUTION = (
    "UK Business Density Map compiled from Business Population Estimates for "
    "the UK and regions (Open Government Licence v3.0). Free to cite with "
    "attribution to Holloway Davies."
)

# Regions/nations to include in the ranked map (excludes the "England" and
# "United Kingdom" aggregate rows, which are shown separately as context).
REGION_ROWS = [
    "North East", "North West", "Yorkshire and the Humber", "East Midlands",
    "West Midlands", "East of England", "London", "South East", "South West",
    "Wales", "Scotland", "Northern Ireland",
]

OUT_PATH = (
    Path(__file__).resolve().parents[2]
    / "generalist" / "web" / "src" / "data" / "uk-business-density-map.json"
)


def fetch_workbook(url: str) -> openpyxl.Workbook:
    r = requests.get(url, timeout=90, headers=UA)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)


def main() -> None:
    print("Fetching Business Population Estimates detailed tables...")
    wb = fetch_workbook(XLSX_URL)

    t8 = list(wb["Table 8"].iter_rows(values_only=True))
    t9 = list(wb["Table 9"].iter_rows(values_only=True))

    density_by_region = {r[0]: r for r in t8 if r[0]}
    counts_by_region = {r[0]: r for r in t9 if r[0]}

    uk_density_row = density_by_region["United Kingdom"]
    uk_counts_row = counts_by_region["United Kingdom"]

    regions = []
    for name in REGION_ROWS:
        d = density_by_region.get(name)
        c = counts_by_region.get(name)
        if d is None or c is None:
            continue
        regions.append({
            "region": name,
            "resident_adults": int(d[1]),
            "density_per_10k_adults": int(d[6]),
            "businesses": int(c[1]),
            "pct_zero_employees": float(c[4]),
            "pct_1_to_49_employees": float(c[5]),
            "pct_50_to_249_employees": float(c[6]),
            "pct_250_plus_employees": float(c[7]),
        })

    regions_ranked = sorted(regions, key=lambda r: -r["density_per_10k_adults"])
    highest = regions_ranked[0]
    lowest = regions_ranked[-1]

    headline = {
        "uk_density_per_10k_adults": int(uk_density_row[6]),
        "uk_total_businesses": int(uk_counts_row[1]),
        "highest_region": highest["region"],
        "highest_density_per_10k_adults": highest["density_per_10k_adults"],
        "lowest_region": lowest["region"],
        "lowest_density_per_10k_adults": lowest["density_per_10k_adults"],
        "density_ratio_highest_to_lowest": round(highest["density_per_10k_adults"] / lowest["density_per_10k_adults"], 2),
        "as_of": "start of 2025",
    }

    snapshot = {
        "meta": {
            "generated_at": date.today().isoformat(),
            "as_of": "start of 2025",
            "sources": [{
                "name": "Business Population Estimates for the UK and regions -- Tables 8 and 9",
                "publisher": PUBLISHER,
                "url": XLSX_URL,
                "release_page": RELEASE_PAGE,
                "licence": LICENCE,
                "retrieved": date.today().isoformat(),
            }],
            "notes": (
                "Density is businesses per 10,000 resident adults (aged 16 and over, mid-2023 "
                "population estimates), from Table 8 of the Business Population Estimates "
                "detailed tables. Business counts and size mix are from Table 9. Figures are "
                "for the whole private sector (companies, sole proprietorships and "
                "partnerships) at the start of 2025. Regional and national totals are "
                "official DBT/ONS statistics, not derived by us."
            ),
            "attribution": ATTRIBUTION,
        },
        "headline": headline,
        "regions": regions_ranked,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n[snapshot] wrote {OUT_PATH}")

    # Self-check
    assert len(regions_ranked) == 12, f"expected 12 regions/nations, got {len(regions_ranked)}"
    assert regions_ranked[0]["density_per_10k_adults"] >= regions_ranked[-1]["density_per_10k_adults"]
    print("[self-check] PASS")
    print(f"UK density: {headline['uk_density_per_10k_adults']}/10k adults")
    print(f"Highest: {highest['region']} ({highest['density_per_10k_adults']}/10k)")
    print(f"Lowest: {lowest['region']} ({lowest['density_per_10k_adults']}/10k)")


if __name__ == "__main__":
    main()
