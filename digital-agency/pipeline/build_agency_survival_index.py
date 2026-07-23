"""UK Agency Survival & Churn Index -- data pipeline (FLAGSHIP asset).

Parses ONS Business Demography 2024 reference tables (Table 5.2a-5.2e:
survival of newly born enterprises by SIC2007 group, one sheet per
birth-year cohort 2019-2023) to build 1-to-5-year survival curves for the
agency SIC cluster.

Isolability, confirmed by inspecting the workbook directly:
  - 731 : Advertising                         -- isolable to 3-digit (includes
                                                  73.11 advertising agencies AND
                                                  73.12 media representation;
                                                  ONS does not split these further)
  - 732 : Market research and public opinion polling -- isolable to 3-digit
  - 741 : Specialised design activities        -- isolable to 3-digit
  - 620 : Computer programming, consultancy and related activities -- this
    is the finest cut ONS publishes for division 62; it already covers both
    62012 (business/domestic software development) and 62020 (IT
    consultancy) with no further split
  - 70  : Activities of head offices; management consultancy activities --
    NOT PR-specific. SIC 70210 (public relations) sits inside 3-digit group
    702 (Management consultancy activities) alongside general management
    consultants, and ONS does not split 702 down to the 4-digit level. This
    group is reported as context only, NOT folded into the combined "agency"
    curve, because management consultancy births vastly outnumber PR births
    and would swamp a genuine agency-cluster figure.

The combined "agency" curve sums raw births and survivor counts for 731 +
732 + 741 + 620 (the four genuinely agency-isolable groups) before dividing,
so it is an exact weighted figure, not an average of percentages -- same
approach as care/pipeline/build_care_survival_index.py's combined SIC 87+88
curve.

Input (already downloaded for the care-sector pilot, reused as-is; this is
the SAME UK-wide ONS release, so cross-niche reuse is exact, not just
similar):
  care/pipeline/raw/ons_business_demography_2024.xlsx

Source: https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable
Licence: Open Government Licence v3.0.

Usage (from repo root):
    python digital-agency/pipeline/build_agency_survival_index.py
"""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
RAW = HERE.parent.parent / "care" / "pipeline" / "raw"
SRC = RAW / "ons_business_demography_2024.xlsx"
OUT = HERE.parent / "web" / "src" / "data" / "uk-agency-survival-churn-index.json"
OUT.parent.mkdir(parents=True, exist_ok=True)

COHORT_SHEETS = [
    ("Table 5.2a", 2019),
    ("Table 5.2b", 2020),
    ("Table 5.2c", 2021),
    ("Table 5.2d", 2022),
    ("Table 5.2e", 2023),
]

# (SIC2007 group code, output key, label, isolable to agency-specific level?)
SEGMENTS = [
    ("731", "advertising", "Advertising (includes media representation, SIC 73.11 & 73.12)", True),
    ("732", "market_research", "Market research and public opinion polling", True),
    ("741", "design", "Specialised design activities", True),
    ("620", "it_consultancy", "Computer programming, consultancy and related activities", True),
    ("70", "management_consultancy_context", "Activities of head offices; management consultancy activities (context only -- NOT PR-specific, see notes)", False),
    ("Total", "all_industries", "All UK industries", True),
]
COMBINE_KEYS = ("advertising", "market_research", "design", "it_consultancy")

HORIZONS = [1, 2, 3, 4, 5]


def _num(v) -> float | None:
    if v in (None, ":", ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_cohort_sheet(ws) -> dict[str, dict]:
    """Return {sic_code: {births, survivors: {h: n|None}, pct: {h: pct|None}}}."""
    out: dict[str, dict] = {}
    codes = {s[0] for s in SEGMENTS}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        label = row[0]
        if not label:
            continue
        raw = str(label).strip()
        code = "Total" if raw == "Total" else raw.split(" ")[0].rstrip(":")
        if code not in codes:
            continue
        births = _num(row[1])
        survivors = {}
        pct = {}
        # columns: 0=label, 1=births, then pairs (survivors,pct) for horizons 1..5
        for h in HORIZONS:
            col = 2 + (h - 1) * 2
            survivors[h] = _num(row[col]) if col < len(row) else None
            pct[h] = _num(row[col + 1]) if col + 1 < len(row) else None
        out[code] = {"births": births, "survivors": survivors, "pct": pct}
    return out


def _stats(row: dict | None) -> dict:
    if not row:
        return {"births": None, **{f"y{h}_count": None for h in HORIZONS}, **{f"y{h}_pct": None for h in HORIZONS}}
    out: dict = {"births": row["births"]}
    for h in HORIZONS:
        out[f"y{h}_count"] = row["survivors"].get(h)
        pct = row["pct"].get(h)
        out[f"y{h}_pct"] = round(pct, 1) if pct is not None else None
    return out


def _combine(rows_by_key: dict[str, dict | None]) -> dict:
    """Sum raw births/survivors across COMBINE_KEYS, then divide -- exact weighted pct."""
    present = [rows_by_key[k] for k in COMBINE_KEYS if rows_by_key.get(k) and rows_by_key[k]["births"] is not None]
    if not present:
        return _stats(None)
    total_births = sum(r["births"] for r in present)
    out: dict = {"births": total_births}
    for h in HORIZONS:
        survivor_vals = [r["survivors"].get(h) for r in present]
        if any(v is None for v in survivor_vals) or total_births == 0:
            out[f"y{h}_count"] = None
            out[f"y{h}_pct"] = None
            continue
        total_survivors = sum(survivor_vals)
        out[f"y{h}_count"] = total_survivors
        out[f"y{h}_pct"] = round(total_survivors / total_births * 100, 1)
    return out


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    rows_by_year: dict[int, dict[str, dict | None]] = {}
    for sheet_name, birth_year in COHORT_SHEETS:
        ws = wb[sheet_name]
        parsed = parse_cohort_sheet(ws)
        rows_by_year[birth_year] = {key: parsed.get(code) for code, key, _label, _iso in SEGMENTS}

    cohorts_out = []
    for _sheet, birth_year in COHORT_SHEETS:
        year_rows = rows_by_year[birth_year]
        cohort: dict = {"birth_year": birth_year}
        for _code, key, _label, _iso in SEGMENTS:
            cohort[key] = _stats(year_rows.get(key))
        cohort["agency_combined"] = _combine(year_rows)
        cohorts_out.append(cohort)

    # Headline: latest full 5-year cohort = 2019 (only one with a full 5yr horizon
    # in this release); latest 1-year cohort = 2023 (most recent published).
    c2019 = next(c for c in cohorts_out if c["birth_year"] == 2019)
    c2023 = next(c for c in cohorts_out if c["birth_year"] == 2023)

    headline = {
        "latest_5yr_cohort_year": 2019,
        "latest_5yr_agency_pct": c2019["agency_combined"]["y5_pct"],
        "latest_5yr_all_industries_pct": c2019["all_industries"]["y5_pct"],
        "latest_5yr_advertising_pct": c2019["advertising"]["y5_pct"],
        "latest_5yr_market_research_pct": c2019["market_research"]["y5_pct"],
        "latest_5yr_design_pct": c2019["design"]["y5_pct"],
        "latest_5yr_it_consultancy_pct": c2019["it_consultancy"]["y5_pct"],
        "latest_1yr_cohort_year": 2023,
        "latest_1yr_agency_pct": c2023["agency_combined"]["y1_pct"],
        "latest_1yr_all_industries_pct": c2023["all_industries"]["y1_pct"],
        "agency_combined_births_2019": c2019["agency_combined"]["births"],
    }

    snapshot = {
        "meta": {
            "title": "UK Agency Survival & Churn Index",
            "description": (
                "How long newly incorporated UK marketing, creative, advertising and digital "
                "agencies actually survive, by 1, 2, 3, 4 and 5 years after formation, built from "
                "ONS Business Demography birth-cohort data for the agency-isolable SIC groups: "
                "731 (advertising, including media representation), 732 (market research), 741 "
                "(specialised design) and 620 (computer programming and IT consultancy)."
            ),
            "generated_at": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "pull_date": dt.date.today().isoformat(),
            "licence": "Open Government Licence v3.0",
            "source": {
                "name": "Business Demography 2024 reference tables (Tables 5.2a-5.2e)",
                "publisher": "Office for National Statistics",
                "url": "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation/datasets/businessdemographyreferencetable",
                "release_date": "2025-11-20",
                "licence": "Open Government Licence v3.0",
            },
            "attribution": (
                "UK Agency Survival & Churn Index data compiled from ONS Business Demography "
                "reference tables (Open Government Licence v3.0). Free to cite with attribution "
                "to Agency Founder Finance."
            ),
            "methodology": (
                "Survival rates are ONS's own published enterprise-level birth-cohort survival "
                "figures (Tables 5.2a-5.2e), read directly from the finest Standard Industrial "
                "Classification 2007 group breakdown ONS publishes for each segment: not "
                "re-derived or estimated. An 'enterprise' birth is a new business appearing on "
                "the Inter-Departmental Business Register in a given year; survival is measured "
                "against VAT/PAYE registration remaining active at each anniversary. Each "
                "birth-year cohort (2019 to 2023) can only show survival up to however many years "
                "have elapsed since that release: the 2019 cohort has the full 5-year curve, the "
                "2023 cohort only 1-year. The combined 'agency' curve sums raw births and survivor "
                "counts for the four agency-isolable groups (731, 732, 741, 620) before dividing, "
                "so it is an exact weighted figure, not an average of the four percentages."
            ),
            "caveats": [
                "Advertising (731) is reported as a single ONS group covering both SIC 73.11 "
                "(advertising agencies) and 73.12 (media representation); ONS does not split "
                "these two 4-digit codes further in the survival tables.",
                "Public relations (SIC 70210) is NOT isolable in this release. It sits inside "
                "3-digit ONS group 702 (management consultancy activities) alongside general "
                "management consultants, who vastly outnumber PR agencies. Group 70 is reported "
                "as context only in this dataset and is deliberately excluded from the combined "
                "'agency' survival curve, because folding it in would swamp the genuine "
                "agency-cluster figure with an unrelated population.",
                "620 (computer programming, consultancy and related activities) is the finest "
                "cut ONS publishes for division 62: it covers both CH SIC 62012 (business and "
                "domestic software development) and 62020 (IT consultancy) with no further "
                "split, and also includes some software agencies that are not marketing or "
                "creative agencies in the everyday sense.",
                "SIC group counts are enterprise-level, not company-level: a single agency "
                "with multiple trading entities or a sole trader/partnership registered for "
                "VAT or PAYE is counted the same way as a limited company, unlike the "
                "Companies House-based UK Agency Formation Index, which counts incorporated "
                "companies specifically.",
                "Counts are control-rounded to base 5 by ONS disclosure control, so small "
                "percentage movements between adjacent cohorts can reflect rounding as much as "
                "real change.",
                "Only the 2019 birth cohort has a complete 5-year curve in this release; later "
                "cohorts (2020 to 2023) are shown only as far as ONS has published, so the "
                "1-year-survival trend across cohorts is the fairest recent-years comparison.",
            ],
        },
        "cohorts": cohorts_out,
        "headline": headline,
        "segments_meta": [
            {"code": code, "key": key, "label": label, "isolable": iso}
            for code, key, label, iso in SEGMENTS
        ],
    }

    OUT.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")
    print(f"[done] wrote {OUT} ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  Headline: {headline}")

    # Self-check
    assert len(cohorts_out) == 5, "expected 5 birth-year cohorts"
    assert headline["latest_5yr_agency_pct"] is not None, "combined agency 5yr pct missing"
    assert 0 < headline["latest_5yr_agency_pct"] < 100, "combined agency 5yr pct out of range"
    assert headline["latest_5yr_all_industries_pct"] is not None, "all-industries 5yr pct missing"
    for c in cohorts_out:
        for key in ("advertising", "market_research", "design", "it_consultancy", "all_industries"):
            if c[key]["births"] is not None:
                assert c[key]["births"] >= 0, f"negative births: {c['birth_year']} {key}"
    print("[self-check] PASS")


if __name__ == "__main__":
    main()
