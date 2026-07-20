"""Build CIS subcontractor invoice templates (XLSX + PDF) into the
construction-cis static downloads dir. Idempotent: overwrites outputs.

Variants: standard-vat, reverse-charge (domestic reverse charge), no-vat.

Run: python scripts/build_cis_invoice_templates.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas

OUT = Path(__file__).resolve().parents[1] / "construction-cis" / "web" / "public" / "downloads"

SLATE = "1E293B"
ORANGE = "F97316"
LIGHT = "F5F5F4"

VARIANTS = {
    "standard-vat": {
        "label": "Standard VAT invoice",
        "vat": "standard",
        "note": "VAT is charged at 20% on the invoice total (labour plus materials). "
                "The CIS deduction is calculated on the labour element excluding VAT.",
    },
    "reverse-charge": {
        "label": "Domestic reverse charge invoice",
        "vat": "reverse",
        "note": "Reverse charge: customer to account for VAT to HMRC. "
                "VAT rate applicable: 20%. No VAT is added to the amount payable. "
                "The CIS deduction is calculated on the labour element excluding VAT.",
    },
    "no-vat": {
        "label": "Non-VAT-registered invoice",
        "vat": "none",
        "note": "No VAT is charged because the supplier is not VAT registered. "
                "The CIS deduction is calculated on the labour element.",
    },
}


def build_xlsx(key: str, spec: dict, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    thin = Side(style="thin", color="D4D4D4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = "£#,##0.00"

    widths = {"A": 42, "B": 14, "C": 8, "D": 12, "E": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def cell(ref, value, *, bold=False, fill=None, fmt=None, size=None,
             color=None, wrap=False, align=None, border=False):
        c = ws[ref]
        c.value = value
        c.font = Font(bold=bold, size=size or 11, color=color or "000000")
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if fmt:
            c.number_format = fmt
        if wrap or align:
            c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical="top")
        if border:
            c.border = box
        return c

    # Header
    ws.merge_cells("A1:E1")
    cell("A1", "INVOICE", bold=True, size=20, color="FFFFFF", fill=SLATE)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:E2")
    cell("A2", f"CIS subcontractor invoice · {spec['label']}", bold=True, size=11, color="FFFFFF", fill=SLATE)

    # From / To
    cell("A4", "From (subcontractor)", bold=True, fill=LIGHT)
    cell("D4", "Invoice details", bold=True, fill=LIGHT)
    from_rows = ["Business name", "Address", "", "UTR number", "VAT number" if spec["vat"] != "none" else "Phone / email"]
    inv_rows = ["Invoice number", "Invoice date", "Payment due", "Contract / site ref", ""]
    for i, (f, d) in enumerate(zip(from_rows, inv_rows), start=5):
        cell(f"A{i}", f)
        cell(f"D{i}", d)
    cell("A11", "To (contractor / customer)", bold=True, fill=LIGHT)
    for i, label in enumerate(["Business name", "Address", "", "VAT number (customer)" if spec["vat"] == "reverse" else ""], start=12):
        if label:
            cell(f"A{i}", label)

    # Line items
    r = 17
    cell(f"A{r}", "Description", bold=True, color="FFFFFF", fill=SLATE, border=True)
    cell(f"B{r}", "Type", bold=True, color="FFFFFF", fill=SLATE, border=True)
    cell(f"C{r}", "Qty", bold=True, color="FFFFFF", fill=SLATE, border=True)
    cell(f"D{r}", "Unit price", bold=True, color="FFFFFF", fill=SLATE, border=True)
    cell(f"E{r}", "Amount", bold=True, color="FFFFFF", fill=SLATE, border=True)

    first_labour, n_labour = r + 1, 4
    first_mat = first_labour + n_labour
    n_mat = 4
    for i in range(n_labour):
        row = first_labour + i
        cell(f"A{row}", "Labour: " if i == 0 else "", border=True)
        cell(f"B{row}", "Labour", border=True)
        cell(f"C{row}", None, border=True)
        cell(f"D{row}", None, fmt=money, border=True)
        cell(f"E{row}", f"=IF(C{row}*D{row}=0,\"\",C{row}*D{row})", fmt=money, border=True)
    for i in range(n_mat):
        row = first_mat + i
        cell(f"A{row}", "Materials: " if i == 0 else "", border=True)
        cell(f"B{row}", "Materials", border=True)
        cell(f"C{row}", None, border=True)
        cell(f"D{row}", None, fmt=money, border=True)
        cell(f"E{row}", f"=IF(C{row}*D{row}=0,\"\",C{row}*D{row})", fmt=money, border=True)

    labour_rng = f"E{first_labour}:E{first_labour + n_labour - 1}"
    mat_rng = f"E{first_mat}:E{first_mat + n_mat - 1}"
    r = first_mat + n_mat + 1

    def total_row(label, formula, *, bold=False, fill=None):
        nonlocal r
        cell(f"D{r}", label, bold=bold, fill=fill, align="right")
        cell(f"E{r}", formula, bold=bold, fill=fill, fmt=money, border=True)
        this = r
        r += 1
        return this

    labour_sub = total_row("Labour subtotal", f"=SUM({labour_rng})")
    total_row("Materials subtotal", f"=SUM({mat_rng})")
    subtotal = total_row("Subtotal (ex VAT)", f"=SUM({labour_rng},{mat_rng})", bold=True)

    # CIS rate cell (editable)
    cell(f"A{r}", "CIS deduction rate (20% registered, 30% unregistered, 0% gross status)", wrap=True)
    rate_ref = f"B{r}"
    c = cell(rate_ref, 0.20, bold=True, fmt="0%", border=True, fill=LIGHT)
    c.alignment = Alignment(horizontal="center")
    cis = total_row("CIS deduction (labour only)", f"=-E{labour_sub}*{rate_ref}", bold=True)

    if spec["vat"] == "standard":
        vat = total_row("VAT @ 20%", f"=E{subtotal}*0.2")
        total_row("Amount payable", f"=E{subtotal}+E{vat}+E{cis}", bold=True, fill="FFEDD5")
    elif spec["vat"] == "reverse":
        total_row("VAT (reverse charge, for information only)", f"=E{subtotal}*0.2")
        total_row("Amount payable", f"=E{subtotal}+E{cis}", bold=True, fill="FFEDD5")
    else:
        total_row("Amount payable", f"=E{subtotal}+E{cis}", bold=True, fill="FFEDD5")

    r += 1
    ws.merge_cells(f"A{r}:E{r + 1}")
    cell(f"A{r}", spec["note"], wrap=True, fill=LIGHT)
    r += 3
    ws.merge_cells(f"A{r}:E{r + 1}")
    cell(f"A{r}", "The CIS deduction applies to the labour element only, excluding VAT. "
                  "Materials you personally purchased for the job are excluded from the deduction base. "
                  "Show your UTR on every CIS invoice.", wrap=True)
    r += 3
    cell(f"A{r}", "Payment details", bold=True, fill=LIGHT)
    for label in ["Account name", "Sort code", "Account number", "Payment reference"]:
        r += 1
        cell(f"A{r}", label)
    r += 2
    cell(f"A{r}", "Template by Trade Tax Specialists · tradetaxspecialists.co.uk/cis-invoice-template", size=9, color="737373")

    wb.save(path)


def build_pdf(key: str, spec: dict, path: Path) -> None:
    c = pdfcanvas.Canvas(str(path), pagesize=A4)
    W, H = A4
    m = 18 * mm
    y = H - m

    def text(s, size=10, bold=False, dy=14, color=colors.black, x=m):
        nonlocal y
        y -= dy
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.setFillColor(color)
        c.drawString(x, y, s)
        c.setFillColor(colors.black)

    def line_field(label, x, w):
        c.setFont("Helvetica", 9)
        c.drawString(x, y, label)
        c.setStrokeColor(colors.Color(0.6, 0.6, 0.6))
        c.line(x + c.stringWidth(label, "Helvetica", 9) + 4, y - 1, x + w, y - 1)

    # Header band
    c.setFillColor(colors.HexColor("#1e293b"))
    c.rect(0, H - 30 * mm, W, 30 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(m, H - 16 * mm, "INVOICE")
    c.setFont("Helvetica", 10)
    c.drawString(m, H - 23 * mm, f"CIS subcontractor invoice · {spec['label']}")
    y = H - 38 * mm

    text("From (subcontractor)", bold=True, dy=0)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(W / 2, y, "Invoice details")
    left = ["Business name:", "Address:", "UTR number:",
            "VAT number:" if spec["vat"] != "none" else "Phone / email:"]
    right = ["Invoice number:", "Invoice date:", "Payment due:", "Contract / site ref:"]
    for lf, rf in zip(left, right):
        y -= 18
        line_field(lf, m, W / 2 - 10 * mm)
        line_field(rf, W / 2, W - m)
    y -= 26
    text("To (contractor / customer)", bold=True, dy=0)
    to_fields = ["Business name:", "Address:"]
    if spec["vat"] == "reverse":
        to_fields.append("Customer VAT number:")
    for lf in to_fields:
        y -= 18
        line_field(lf, m, W / 2 - 10 * mm)

    # Items table
    y -= 30
    rows = 8
    row_h = 18
    x_cols = [m, m + 78 * mm, m + 108 * mm, m + 122 * mm, m + 146 * mm, W - m]
    c.setFillColor(colors.HexColor("#1e293b"))
    c.rect(m, y - 4, W - 2 * m, row_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    for label, x in zip(["Description", "Type", "Qty", "Unit price", "Amount"], x_cols):
        c.drawString(x + 2, y, label)
    c.setFillColor(colors.black)
    c.setStrokeColor(colors.Color(0.8, 0.8, 0.8))
    c.setFont("Helvetica", 9)
    for i in range(rows):
        yy = y - (i + 1) * row_h
        c.rect(m, yy - 4, W - 2 * m, row_h, fill=0, stroke=1)
        c.drawString(x_cols[1] + 2, yy, "Labour" if i < 4 else "Materials")
    for x in x_cols[1:-1]:
        c.line(x, y - 4 + row_h, x, y - 4 - rows * row_h)
    y -= (rows + 1) * row_h + 6

    totals = ["Labour subtotal", "Materials subtotal", "Subtotal (ex VAT)",
              "CIS deduction (labour only, at 20% / 30%)"]
    if spec["vat"] == "standard":
        totals += ["VAT @ 20%"]
    elif spec["vat"] == "reverse":
        totals += ["VAT (reverse charge, for information only)"]
    totals += ["Amount payable"]
    for t in totals:
        y -= 18
        c.setFont("Helvetica-Bold" if t in ("Subtotal (ex VAT)", "Amount payable") else "Helvetica", 9)
        c.drawRightString(x_cols[4] - 4, y, t)
        c.setStrokeColor(colors.Color(0.6, 0.6, 0.6))
        c.line(x_cols[4], y - 1, W - m, y - 1)

    y -= 10
    text(spec["note"], size=9, dy=20)
    text("The CIS deduction applies to the labour element only, excluding VAT. Materials you personally", size=9, dy=14)
    text("purchased for the job are excluded from the deduction base. Show your UTR on every CIS invoice.", size=9, dy=12)

    y -= 8
    text("Payment details", bold=True, dy=16)
    y -= 4
    for lf in ["Account name:", "Sort code:", "Account number:", "Payment reference:"]:
        y -= 16
        line_field(lf, m, W / 2 - 10 * mm)

    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.Color(0.45, 0.45, 0.45))
    c.drawString(m, 12 * mm, "Template by Trade Tax Specialists · tradetaxspecialists.co.uk/cis-invoice-template")
    c.save()


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    for key, spec in VARIANTS.items():
        xlsx = OUT / f"cis-subcontractor-invoice-template-{key}.xlsx"
        pdf = OUT / f"cis-subcontractor-invoice-template-{key}.pdf"
        build_xlsx(key, spec, xlsx)
        build_pdf(key, spec, pdf)
        print(f"built {xlsx.name} ({xlsx.stat().st_size} B), {pdf.name} ({pdf.stat().st_size} B)")

    # self-check: reverse-charge wording present, CIS formula references labour subtotal
    from openpyxl import load_workbook
    wb = load_workbook(OUT / "cis-subcontractor-invoice-template-reverse-charge.xlsx")
    vals = [str(c.value) for row in wb.active.iter_rows() for c in row if c.value]
    assert any("Reverse charge: customer to account for VAT to HMRC" in v for v in vals)
    assert any(v.startswith("=-E") for v in vals), "CIS deduction formula missing"
    print("self-check OK")


if __name__ == "__main__":
    main()
