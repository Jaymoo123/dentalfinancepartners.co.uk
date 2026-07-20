"""Generate the CIS Payment & Deduction Statement template (XLSX + PDF)
into construction-cis/web/public/downloads/. Re-runnable.
Fields per CIS Regulation 4(8): contractor name + employer ref, tax month
ended 5th, subcontractor name/UTR/verification no., gross amount (excl VAT),
cost of materials, amount deducted, rate.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

OUT = Path(__file__).resolve().parents[1] / "construction-cis" / "web" / "public" / "downloads"
OUT.mkdir(parents=True, exist_ok=True)

INK = "1E293B"
ORANGE = "F97316"

CONTRACTOR_FIELDS = [
    ("Contractor name", ""),
    ("Contractor address", ""),
    ("Employer's PAYE reference", "e.g. 123/AB45678"),
    ("Tax month ended 5th of", "e.g. 5 August 2026"),
]
SUB_FIELDS = [
    ("Subcontractor name", ""),
    ("Unique Taxpayer Reference (UTR)", "10 digits"),
    ("Verification number", "Only required if deduction made at the 30% rate, e.g. V1234567890A"),
]
PAY_ROWS = [
    ("Gross amount paid (excluding VAT)", "£", "Total payment for the month before any deduction, excluding VAT"),
    ("Less: cost of materials", "£", "Direct cost of materials supplied by the subcontractor"),
    ("Amount liable to deduction", "£", "Gross amount minus cost of materials"),
    ("Deduction rate applied", "20% / 30%", "20% registered, 30% unverified or unregistered"),
    ("Amount deducted", "£", "Amount liable to deduction x rate"),
    ("Amount payable to subcontractor", "£", "Gross amount minus amount deducted"),
]
FOOTNOTE = (
    "Issued under Regulation 4(8) of the Income Tax (Construction Industry Scheme) "
    "Regulations 2005. Contractors must give this statement to every subcontractor "
    "paid under deduction within 14 days of the end of the tax month (by the 19th). "
    "Subcontractors: keep this statement. You need it to claim credit for the "
    "deduction on your Self Assessment return or company EPS."
)
CREDIT = "Free template from Trade Tax Specialists · tradetaxspecialists.co.uk"


def build_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CIS Statement"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 52
    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor=INK)
    sub_fill = PatternFill("solid", fgColor="F1F5F9")
    r = 1
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(r, 1, "CIS Payment and Deduction Statement")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = head_fill
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(r, 1, "Regulation 4(8), Income Tax (Construction Industry Scheme) Regulations 2005")
    c.font = Font(size=9, italic=True, color="64748B")
    r += 2

    def section(title, rows):
        nonlocal r
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(r, 1, title)
        c.font = Font(bold=True, size=11, color=INK)
        c.fill = sub_fill
        c.border = box
        r += 1
        for label, note in rows:
            ws.cell(r, 1, label).font = Font(bold=True, size=10)
            ws.cell(r, 1).border = box
            ws.cell(r, 2).border = box  # blank fill-in cell
            ws.cell(r, 3, note).font = Font(size=9, italic=True, color="64748B")
            ws.cell(r, 3).border = box
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

    section("Contractor details", CONTRACTOR_FIELDS)
    section("Subcontractor details", SUB_FIELDS)

    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(r, 1, "Payment and deduction")
    c.font = Font(bold=True, size=11, color=INK)
    c.fill = sub_fill
    c.border = box
    r += 1
    first_amount_row = None
    for label, unit, note in PAY_ROWS:
        ws.cell(r, 1, label).font = Font(bold=True, size=10)
        ws.cell(r, 1).border = box
        b = ws.cell(r, 2)
        b.border = box
        if unit == "£":
            b.number_format = '£#,##0.00'
            if first_amount_row is None:
                first_amount_row = r
        ws.cell(r, 3, note).font = Font(size=9, italic=True, color="64748B")
        ws.cell(r, 3).border = box
        ws.row_dimensions[r].height = 20
        r += 1
    # formulas: liable = gross - materials; payable = gross - deducted
    g, m, liable, rate, ded, pay = range(first_amount_row, first_amount_row + 6)
    ws.cell(liable, 2).value = f"=IF(B{g}=\"\",\"\",B{g}-B{m})"
    ws.cell(pay, 2).value = f"=IF(OR(B{g}=\"\",B{ded}=\"\"),\"\",B{g}-B{ded})"
    r += 1

    ws.merge_cells(f"A{r}:C{r + 2}")
    c = ws.cell(r, 1, FOOTNOTE)
    c.font = Font(size=9, color="475569")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 3
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r, 1, CREDIT).font = Font(size=8, color="94A3B8")
    wb.save(path)


def build_pdf(path: Path):
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#" + INK), alignment=0, spaceAfter=2)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"))
    sect = ParagraphStyle("sect", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#" + INK), spaceBefore=10, spaceAfter=4, fontName="Helvetica-Bold")
    doc = SimpleDocTemplate(str(path), pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    story = [
        Paragraph("CIS Payment and Deduction Statement", h1),
        Paragraph("Regulation 4(8), Income Tax (Construction Industry Scheme) Regulations 2005", small),
        Spacer(1, 6),
    ]
    ts = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#64748B")),
        ("FONTSIZE", (2, 0), (2, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    widths = [62 * mm, 42 * mm, 70 * mm]
    note_style = ParagraphStyle("note", parent=small, fontSize=7.5)

    def table(title, rows):
        story.append(Paragraph(title, sect))
        data = [[label, "", Paragraph(note, note_style)] for label, note in rows]
        t = Table(data, colWidths=widths)
        t.setStyle(ts)
        story.append(t)

    table("Contractor details", CONTRACTOR_FIELDS)
    table("Subcontractor details", SUB_FIELDS)
    story.append(Paragraph("Payment and deduction", sect))
    data = [[label, unit if unit != "£" else "£", Paragraph(note, note_style)] for label, unit, note in PAY_ROWS]
    t = Table(data, colWidths=widths)
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 12))
    story.append(Paragraph(FOOTNOTE, small))
    story.append(Spacer(1, 6))
    story.append(Paragraph(CREDIT, small))
    doc.build(story)


if __name__ == "__main__":
    xlsx = OUT / "cis-payment-deduction-statement-template.xlsx"
    pdf = OUT / "cis-payment-deduction-statement-template.pdf"
    build_xlsx(xlsx)
    build_pdf(pdf)
    for p in (xlsx, pdf):
        assert p.exists() and p.stat().st_size > 2000, p
        print(p, p.stat().st_size, "bytes")
